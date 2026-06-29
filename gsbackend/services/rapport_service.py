# ============================================================================
#  services/rapport_service.py  —  VERSION COMPLÈTE (helpers v2 intégrés)
# ----------------------------------------------------------------------------
#  Ce fichier remplace intégralement l'ancien rapport_service.py.
#
#  CE QUI A CHANGÉ (uniquement la zone helpers, en haut) :
#   • Graphiques  : barres en couleur unique, camemberts → donuts (total au
#                   centre, % masqués <4 %), courbes avec aire, grille discrète,
#                   DPI 150→200, police propre.
#   • PDF         : KPI en grille de CARTES, pagination « Page X / Y » sur
#                   chaque page (canvas 2 passes), filets de tableau allégés.
#   • Excel       : nombres stockés comme vrais nombres (format milliers),
#                   excel_autofit robuste aux cellules fusionnées, option freeze.
#   • Nouveaux    : make_histogram() et make_grouped_bar_chart() (pour insights).
#
#  Les 12 générateurs de RapportGeneratorService sont INCHANGÉS : ils héritent
#  automatiquement des améliorations car ils passent tous par ces helpers.
#
#  NB — 2 bugs connus restent dans les générateurs (non corrigés ici, sur
#  demande) :
#    1. generer_synthese_mensuelle : l'évolution du taux d'acceptation
#       (ex. "85.5%") plante sur int() et s'affiche "—" → passer en float().
#    2. generer_audit_actions : par_user / par_mois_data calculés sur qs[:1000]
#       alors que total = qs.count() → faux au-delà de 1000 actions
#       → utiliser une agrégation SQL.
# ============================================================================

import os
from io import BytesIO
from datetime import date, timedelta
from django.utils import timezone
from django.db.models import Q, Count, Avg, F

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker

from reportlab.pdfgen import canvas as _rl_canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm, inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, Image as RLImage, KeepTogether, PageBreak,
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint

from stages.models import UserAction, Stagiaire, Demande, DemandeAttestation, ConventionStage
from utilisateurs.models import Utilisateur


# ── Constantes ──────────────────────────────────────────────────────────────

LOGO_PATH = os.path.join(os.path.dirname(__file__), 'assets', 'ceb-logo.png')

# Palette professionnelle
CLR_PRIMARY    = '#1e293b'
CLR_ACCENT     = '#6366f1'
CLR_SUCCESS    = '#10b981'
CLR_DANGER     = '#ef4444'
CLR_WARNING    = '#f59e0b'
CLR_INFO       = '#3b82f6'
CLR_LIGHT_BG   = '#f8fafc'
CLR_BORDER     = '#e2e8f0'
CLR_TEXT_MUTED = '#64748b'
CLR_TEXT_LIGHT = '#94a3b8'

HEADER_COLOR  = colors.HexColor(CLR_PRIMARY)
ACCENT_COLOR  = colors.HexColor(CLR_ACCENT)
SUCCESS_COLOR = colors.HexColor(CLR_SUCCESS)
DANGER_COLOR  = colors.HexColor(CLR_DANGER)
WARNING_COLOR = colors.HexColor(CLR_WARNING)
INFO_COLOR    = colors.HexColor(CLR_INFO)
LIGHT_BG      = colors.HexColor(CLR_LIGHT_BG)
BORDER_COLOR  = colors.HexColor(CLR_BORDER)

# Palette pour graphiques matplotlib
CHART_COLORS = ['#6366f1', '#10b981', '#f59e0b', '#ef4444', '#3b82f6',
                '#8b5cf6', '#ec4899', '#14b8a6', '#f97316', '#06b6d4']
CHART_BG  = '#ffffff'
CHART_DPI = 200          # ⬆ 150 → 200 : graphiques plus nets dans le PDF

NOMS_MOIS_COURT = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun', 'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']
NOMS_MOIS = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
             'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
JOURS_LABELS = ['Dim', 'Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam']


# ── Helpers Graphiques (matplotlib) ─────────────────────────────────────────

def _setup_chart_style():
    """Style global matplotlib : grille discrète, spines épurés, police propre."""
    plt.rcParams.update({
        'figure.facecolor': CHART_BG,
        'axes.facecolor':    CHART_BG,
        'axes.edgecolor':    CLR_BORDER,
        'axes.linewidth':    0.8,
        'axes.labelcolor':   CLR_PRIMARY,
        'axes.titlesize':    11,
        'axes.titleweight':  'bold',
        'axes.titlepad':     14,
        'axes.spines.top':   False,
        'axes.spines.right': False,
        'axes.grid':         True,
        'grid.alpha':        0.25,
        'grid.color':        CLR_BORDER,
        'grid.linewidth':    0.6,
        'xtick.color':       CLR_TEXT_MUTED,
        'ytick.color':       CLR_TEXT_MUTED,
        'xtick.labelsize':   8.5,
        'ytick.labelsize':   8.5,
        'font.size':         9,
        'font.family':       'sans-serif',
        'font.sans-serif':   ['DejaVu Sans', 'Arial', 'Helvetica', 'sans-serif'],
    })


def _chart_to_image(fig, width=16*cm, height=8*cm):
    """Convertit une figure matplotlib en Image ReportLab (haute résolution)."""
    buf = BytesIO()
    fig.savefig(buf, format='png', dpi=CHART_DPI, bbox_inches='tight',
                facecolor=CHART_BG, edgecolor='none')
    plt.close(fig)
    buf.seek(0)
    return RLImage(buf, width=width, height=height)


def make_bar_chart(labels, values, title='', color=CLR_ACCENT, horizontal=False,
                   width=16*cm, height=8*cm, value_labels=True):
    """Bar chart épuré. Couleur unique par défaut (plus lisible qu'un arc-en-ciel)."""
    _setup_chart_style()
    fig, ax = plt.subplots(figsize=(width / cm * 0.35, height / cm * 0.35))
    vmax = max(values) if values and max(values) > 0 else 1

    if horizontal:
        bars = ax.barh(labels, values, color=color, edgecolor='white',
                       height=0.62, linewidth=0.5, zorder=3)
        ax.set_xlim(0, vmax * 1.18)
        ax.invert_yaxis()
        ax.grid(axis='x', alpha=0.25)
        ax.grid(axis='y', visible=False)
        ax.xaxis.set_major_locator(ticker.MaxNLocator(integer=True))
        if value_labels:
            for b, v in zip(bars, values):
                ax.text(b.get_width() + vmax * 0.015, b.get_y() + b.get_height() / 2,
                        str(v), va='center', fontsize=8, fontweight='bold', color=CLR_PRIMARY)
    else:
        bars = ax.bar(labels, values, color=color, edgecolor='white',
                      width=0.62, linewidth=0.5, zorder=3)
        ax.set_ylim(0, vmax * 1.22)
        ax.grid(axis='y', alpha=0.25)
        ax.grid(axis='x', visible=False)
        ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
        if value_labels:
            for b, v in zip(bars, values):
                ax.text(b.get_x() + b.get_width() / 2, b.get_height() + vmax * 0.015,
                        str(v), ha='center', fontsize=8, fontweight='bold', color=CLR_PRIMARY)
        if labels and max((len(str(l)) for l in labels), default=0) > 6 and len(labels) > 5:
            plt.xticks(rotation=35, ha='right')

    if title:
        ax.set_title(title, color=CLR_PRIMARY)
    plt.tight_layout()
    return _chart_to_image(fig, width, height)


def make_pie_chart(labels, values, title='', width=10*cm, height=10*cm):
    """Donut moderne : total au centre, % masqués sous 4 %, légende à droite."""
    _setup_chart_style()
    fig, ax = plt.subplots(figsize=(width / cm * 0.35, height / cm * 0.35))

    filtered = [(l, v) for l, v in zip(labels, values) if v and v > 0]
    if not filtered:
        ax.text(0.5, 0.5, 'Aucune donnée', ha='center', va='center',
                fontsize=11, color=CLR_TEXT_MUTED, transform=ax.transAxes)
        ax.axis('off')
        if title:
            ax.set_title(title, color=CLR_PRIMARY)
        plt.tight_layout()
        return _chart_to_image(fig, width, height)

    f_labels, f_values = zip(*filtered)
    total = sum(f_values)
    chart_colors = CHART_COLORS[:len(f_labels)]

    def _auto(pct):
        return f'{pct:.0f}%' if pct >= 4 else ''

    wedges, texts, autotexts = ax.pie(
        f_values, labels=None, autopct=_auto, colors=chart_colors, startangle=90,
        wedgeprops={'edgecolor': 'white', 'linewidth': 2, 'width': 0.42},
        pctdistance=0.78,
    )
    for t in autotexts:
        t.set_fontsize(8)
        t.set_fontweight('bold')
        t.set_color('white')

    ax.text(0, 0.04, str(total), ha='center', va='center',
            fontsize=15, fontweight='bold', color=CLR_PRIMARY)
    ax.text(0, -0.16, 'total', ha='center', va='center',
            fontsize=7.5, color=CLR_TEXT_MUTED)

    ax.legend(f_labels, loc='center left', bbox_to_anchor=(1, 0.5),
              frameon=False, fontsize=8)
    ax.set(aspect='equal')
    if title:
        ax.set_title(title, color=CLR_PRIMARY)
    plt.tight_layout()
    return _chart_to_image(fig, width, height)


def make_line_chart(labels, datasets, title='', width=16*cm, height=8*cm):
    """Line chart. Série unique : aire + valeurs. Séries multiples : dernière
    valeur étiquetée seulement. datasets = [(name, values), ...]"""
    _setup_chart_style()
    fig, ax = plt.subplots(figsize=(width / cm * 0.35, height / cm * 0.35))

    x = list(range(len(labels)))
    single = len(datasets) == 1

    for i, (name, values) in enumerate(datasets):
        color = CHART_COLORS[i % len(CHART_COLORS)]
        ax.plot(x, values, marker='o', markersize=4, linewidth=2,
                color=color, label=name, zorder=3)
        if single:
            ax.fill_between(x, values, alpha=0.12, color=color, zorder=1)
            for xi, y in zip(x, values):
                ax.annotate(str(y), (xi, y), textcoords="offset points",
                            xytext=(0, 8), ha='center', fontsize=7,
                            color=color, fontweight='bold')
        elif values:
            ax.annotate(str(values[-1]), (x[-1], values[-1]),
                        textcoords="offset points", xytext=(5, 4),
                        fontsize=7, color=color, fontweight='bold')

    ax.set_ylim(0, None)
    ax.set_xticks(x)
    ax.set_xticklabels(labels,
                       rotation=45 if len(labels) > 8 else 0,
                       ha='right' if len(labels) > 8 else 'center')
    ax.grid(axis='y', alpha=0.25)
    ax.grid(axis='x', visible=False)
    ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
    if len(datasets) > 1:
        ax.legend(frameon=False, fontsize=8)
    if title:
        ax.set_title(title, color=CLR_PRIMARY)
    plt.tight_layout()
    return _chart_to_image(fig, width, height)


def make_histogram(values, title='', xlabel='', bins=12, color=CLR_INFO,
                   width=16*cm, height=8*cm):
    """Histogramme de distribution (délais, scores IA, durées de stage…)."""
    _setup_chart_style()
    fig, ax = plt.subplots(figsize=(width / cm * 0.35, height / cm * 0.35))
    vals = [v for v in values if v is not None]

    if not vals:
        ax.text(0.5, 0.5, 'Aucune donnée', ha='center', va='center',
                fontsize=11, color=CLR_TEXT_MUTED, transform=ax.transAxes)
        ax.axis('off')
    else:
        ax.hist(vals, bins=bins, color=color, edgecolor='white', linewidth=0.6, zorder=3)
        ax.grid(axis='y', alpha=0.25)
        ax.grid(axis='x', visible=False)
        if xlabel:
            ax.set_xlabel(xlabel)
        ax.set_ylabel('Effectif')
        ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))

    if title:
        ax.set_title(title, color=CLR_PRIMARY)
    plt.tight_layout()
    return _chart_to_image(fig, width, height)


def make_grouped_bar_chart(labels, series, title='', width=16*cm, height=8*cm,
                           value_labels=True):
    """Barres groupées génériques. series = [(nom, valeurs), ...]."""
    _setup_chart_style()
    fig, ax = plt.subplots(figsize=(width / cm * 0.35, height / cm * 0.35))

    n = max(len(series), 1)
    x = list(range(len(labels)))
    total_w = 0.8
    bw = total_w / n

    for i, (name, values) in enumerate(series):
        color = CHART_COLORS[i % len(CHART_COLORS)]
        offsets = [xi - total_w / 2 + bw / 2 + i * bw for xi in x]
        bars = ax.bar(offsets, values, bw, label=name, color=color,
                      edgecolor='white', linewidth=0.5, zorder=3)
        if value_labels:
            for b, v in zip(bars, values):
                ax.text(b.get_x() + b.get_width() / 2, b.get_height(),
                        str(v), ha='center', va='bottom', fontsize=6.5, color=CLR_PRIMARY)

    ax.set_xticks(x)
    ax.set_xticklabels(labels,
                       rotation=45 if len(labels) > 8 else 0,
                       ha='right' if len(labels) > 8 else 'center')
    ax.grid(axis='y', alpha=0.25)
    ax.grid(axis='x', visible=False)
    ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
    if n > 1:
        ax.legend(frameon=False, fontsize=8)
    if title:
        ax.set_title(title, color=CLR_PRIMARY)
    plt.tight_layout()
    return _chart_to_image(fig, width, height)


# ── Helpers PDF ──────────────────────────────────────────────────────────────

class NumberedCanvas(_rl_canvas.Canvas):
    """Canvas 2 passes : pied de page « Page X / Y » + filet sur chaque page."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self._draw_footer(total)
            super().showPage()
        super().save()

    def _draw_footer(self, total):
        self.saveState()
        self.setStrokeColor(BORDER_COLOR)
        self.setLineWidth(0.5)
        self.line(2 * cm, 1.5 * cm, A4[0] - 2 * cm, 1.5 * cm)
        self.setFont('Helvetica', 7)
        self.setFillColor(colors.HexColor(CLR_TEXT_LIGHT))
        self.drawString(2 * cm, 1.05 * cm, "CEB · Gestion des Stages")
        self.drawRightString(A4[0] - 2 * cm, 1.05 * cm, f"Page {self._pageNumber} / {total}")
        self.restoreState()


class NumberedDocTemplate(SimpleDocTemplate):
    """Injecte NumberedCanvas automatiquement : doc.build(elements) suffit."""

    def build(self, flowables, **kwargs):
        kwargs.setdefault('canvasmaker', NumberedCanvas)
        return super().build(flowables, **kwargs)


def make_pdf_doc(buffer, title):
    """Document A4 paginé (marges standards)."""
    return NumberedDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
        title=title,
    )


def pdf_header(title, subtitle, styles):
    """Bloc d'en-tête avec logo CEB, titre et sous-titre."""
    elements = []

    if os.path.exists(LOGO_PATH):
        logo = RLImage(LOGO_PATH, width=1.5*cm, height=1.5*cm)
        org_text = Paragraph(
            '<font color="#1e293b"><b>CEB</b></font><br/>'
            '<font size="7" color="#64748b">Communauté Électrique du Bénin</font><br/>'
            '<font size="6" color="#94a3b8">Gestion des Stages</font>',
            ParagraphStyle('orgHeader', fontSize=10, leading=13)
        )
        date_text = Paragraph(
            f'<font size="7" color="#94a3b8">Généré le {date.today().strftime("%d/%m/%Y")} '
            f'à {timezone.now().strftime("%H:%M")}</font>',
            ParagraphStyle('dateRight', fontSize=7, alignment=TA_RIGHT)
        )
        header_table = Table([[logo, org_text, date_text]], colWidths=[2*cm, 10*cm, 5*cm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(header_table)
    else:
        elements.append(Paragraph(
            '<font color="#6366f1"><b>CEB</b></font> — Gestion des Stages',
            ParagraphStyle('org', fontSize=9, textColor=colors.HexColor(CLR_TEXT_MUTED), alignment=TA_LEFT)
        ))

    elements.append(Spacer(1, 0.4*cm))
    elements.append(HRFlowable(width="100%", thickness=2, color=ACCENT_COLOR, spaceAfter=8))
    elements.append(Paragraph(
        f'<b>{title}</b>',
        ParagraphStyle('title', fontSize=16, textColor=HEADER_COLOR, spaceAfter=4, leading=20)
    ))
    if subtitle:
        elements.append(Paragraph(
            subtitle,
            ParagraphStyle('sub', fontSize=10, textColor=colors.HexColor(CLR_TEXT_MUTED), spaceAfter=6)
        ))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=BORDER_COLOR, spaceAfter=16))
    return elements


def pdf_section_title(text, styles, color=CLR_ACCENT):
    """Titre de section avec puce de couleur."""
    return [
        Spacer(1, 10),
        Paragraph(
            f'<font color="{color}">■</font>  <b>{text}</b>',
            ParagraphStyle('sectionTitle', fontSize=12, textColor=HEADER_COLOR, spaceAfter=8, leading=16)
        ),
    ]


def _kpi_card(label, value, accent, card_w):
    """Une carte KPI : grand chiffre coloré + libellé, liseré de couleur à gauche."""
    val_str = str(value)
    n = len(val_str)
    val_size = 20 if n <= 6 else 15 if n <= 11 else 10
    val_p = Paragraph(
        f'<b>{val_str}</b>',
        ParagraphStyle('kv', fontName='Helvetica-Bold', fontSize=val_size,
                       leading=val_size + 2, textColor=colors.HexColor(accent))
    )
    lbl_p = Paragraph(
        label.upper(),
        ParagraphStyle('kl', fontName='Helvetica', fontSize=7, leading=9,
                       textColor=colors.HexColor(CLR_TEXT_MUTED))
    )
    card = Table([[val_p], [lbl_p]], colWidths=[card_w])
    card.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), LIGHT_BG),
        ('LINEBEFORE',    (0, 0), (0, -1), 3, colors.HexColor(accent)),
        ('BOX',           (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ('ROUNDEDCORNERS', [5, 5, 5, 5]),
        ('LEFTPADDING',   (0, 0), (-1, -1), 10),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
        ('TOPPADDING',    (0, 0), (0, 0), 9),
        ('BOTTOMPADDING', (0, 0), (0, 0), 1),
        ('TOPPADDING',    (0, 1), (0, 1), 0),
        ('BOTTOMPADDING', (0, 1), (0, 1), 9),
    ]))
    return card


def pdf_kpi_table(kpis, col_widths=None):
    """⬆ Grille de CARTES KPI (drop-in du tableau 2 colonnes).
    kpis = [(label, value[, color]), ...]. 4 cartes par ligne, enroulement auto."""
    if not kpis:
        return Spacer(1, 1)

    per_row = 4
    avail = 17 * cm
    gap = 0.35 * cm
    card_w = (avail - (per_row - 1) * gap) / per_row

    cards = [
        _kpi_card(k[0], k[1], (k[2] if len(k) > 2 and k[2] else CLR_ACCENT), card_w)
        for k in kpis
    ]

    cw = []
    for i in range(per_row):
        cw.append(card_w)
        if i < per_row - 1:
            cw.append(gap)

    rows = []
    for i in range(0, len(cards), per_row):
        group = cards[i:i + per_row]
        line = []
        for j in range(per_row):
            line.append(group[j] if j < len(group) else '')
            if j < per_row - 1:
                line.append('')
        rows.append(line)
        rows.append([Spacer(1, 0.3 * cm)] + [''] * (per_row * 2 - 2))
    if rows:
        rows.pop()

    t = Table(rows, colWidths=cw)
    t.setStyle(TableStyle([
        ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING',   (0, 0), (-1, -1), 0),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
        ('TOPPADDING',    (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    return t


def pdf_table(data, col_widths, header_color=None):
    """Table ReportLab avec style standard amélioré (filets horizontaux seuls)."""
    hc = header_color or ACCENT_COLOR
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), hc),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 8),
        ('ALIGN',         (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING',    (0, 0), (-1, 0), 8),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_BG]),
        ('ALIGN',         (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
        ('TOPPADDING',    (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 7),
        ('LINEBELOW',     (0, 1), (-1, -1), 0.3, BORDER_COLOR),
        ('LINEBELOW',     (0, 0), (-1, 0), 1.5, hc),
    ]))
    return t


def pdf_footer_note(text):
    """Note de bas de contenu (en plus de la pagination de page)."""
    return Paragraph(
        f'<font size="7" color="#94a3b8"><i>{text}</i></font>',
        ParagraphStyle('footer', fontSize=7, textColor=colors.HexColor(CLR_TEXT_LIGHT), spaceBefore=20)
    )


# ── Helpers Excel ─────────────────────────────────────────────────────────────

EXCEL_BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0'),
)


def make_excel_wb(title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]
    ws.sheet_properties.tabColor = '6366F1'
    return wb, ws


def excel_title_block(ws, title, subtitle='', merge_cols=5):
    """Bloc titre professionnel en haut de la feuille."""
    end_col = get_column_letter(merge_cols)

    ws.merge_cells(f'A1:{end_col}1')
    cell = ws['A1']
    cell.value = title
    cell.font = Font(bold=True, size=14, color='1E293B')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill('solid', fgColor='F1F5F9')
    ws.row_dimensions[1].height = 36

    if subtitle:
        ws.merge_cells(f'A2:{end_col}2')
        cell2 = ws['A2']
        cell2.value = subtitle
        cell2.font = Font(size=9, italic=True, color='64748B')
        cell2.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 22
        return 4
    return 3


def excel_header_row(ws, headers, row=1, fill_color="1e293b", freeze=False):
    """Ligne d'en-tête stylée. freeze=True fige les volets sous l'en-tête."""
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = EXCEL_BORDER
    ws.row_dimensions[row].height = 28
    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=1)


def excel_data_row(ws, values, row, even=False):
    """Ligne de données. ⬆ Nombres stockés comme vrais nombres (format milliers)."""
    fill = PatternFill("solid", fgColor="F8FAFC") if even else PatternFill("solid", fgColor="FFFFFF")
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.fill = fill
        cell.border = EXCEL_BORDER
        if isinstance(v, bool):
            pass
        elif isinstance(v, int):
            cell.number_format = '#,##0'
        elif isinstance(v, float):
            cell.number_format = '#,##0.0'
    ws.row_dimensions[row].height = 20


def excel_autofit(ws):
    """Ajuste les largeurs. ⬆ Robuste aux cellules fusionnées (blocs titre)."""
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c, MergedCell) or c.value is None:
                continue
            col = c.column_letter
            widths[col] = max(widths.get(col, 0), len(str(c.value)))
    for col, w in widths.items():
        ws.column_dimensions[col].width = min(w + 4, 50)


def excel_add_bar_chart(ws, title, categories_ref, values_ref, anchor='E2',
                        width=18, height=12, style=10):
    """Ajoute un bar chart dans une feuille Excel."""
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.style = style
    chart.y_axis.title = "Nombre"
    chart.x_axis.title = None
    chart.width = width
    chart.height = height
    chart.legend = None

    data = Reference(ws, **values_ref)
    cats = Reference(ws, **categories_ref)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4

    series = chart.series[0]
    series.graphicalProperties.solidFill = '6366F1'

    ws.add_chart(chart, anchor)
    return chart


def excel_add_pie_chart(ws, title, categories_ref, values_ref, anchor='E2',
                        width=14, height=12):
    """Ajoute un pie chart dans une feuille Excel."""
    chart = PieChart()
    chart.title = title
    chart.width = width
    chart.height = height

    data = Reference(ws, **values_ref)
    cats = Reference(ws, **categories_ref)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = False
    chart.dataLabels.showCatName = True

    pie_colors = ['6366F1', '10B981', 'F59E0B', 'EF4444', '3B82F6', '8B5CF6', 'EC4899', '14B8A6']
    for i in range(min(len(pie_colors), 8)):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = pie_colors[i]
        chart.series[0].data_points.append(pt)

    ws.add_chart(chart, anchor)
    return chart


def excel_add_line_chart(ws, title, categories_ref, values_refs, anchor='E2',
                         width=18, height=12):
    """Ajoute un line chart. values_refs = liste de dicts Reference."""
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height
    chart.y_axis.title = "Nombre"

    cats = Reference(ws, **categories_ref)
    for vr in values_refs:
        data = Reference(ws, **vr)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    line_colors = ['6366F1', '10B981', 'F59E0B', 'EF4444']
    for i, s in enumerate(chart.series):
        s.graphicalProperties.line.solidFill = line_colors[i % len(line_colors)]
        s.graphicalProperties.line.width = 25000
        s.smooth = True

    ws.add_chart(chart, anchor)
    return chart


# ── Service principal ─────────────────────────────────────────────────────────

class RapportGeneratorService:

    # ─────────────────────────────────────────────────────────────────────────
    # OPÉRATIONNEL
    # ─────────────────────────────────────────────────────────────────────────

    @staticmethod
    def generer_demandes_en_cours(params, format='pdf'):
        statut       = params.get('statut', 'tous')
        seuil_jours  = int(params.get('seuil_jours', 14))

        qs = Demande.objects.filter(est_archivee=False).exclude(statut_stage__in=["Acceptée", "Refusée"])
        if statut != 'tous':
            qs = qs.filter(statut_stage=statut)
        qs = qs.select_related('etablissement').order_by('date_soumission')

        today = date.today()
        rows = []
        statut_counts = {}
        for d in qs:
            delta = (today - d.date_soumission.date()).days if d.date_soumission else 0
            s = d.statut_stage or 'N/A'
            statut_counts[s] = statut_counts.get(s, 0) + 1
            rows.append({
                'tracking_id':   d.tracking_id,
                'nom':           f"{d.etudiant_prenom} {d.etudiant_nom}",
                'statut':        s,
                'date_soumission': d.date_soumission.strftime('%d/%m/%Y') if d.date_soumission else '',
                'jours_ecoules': delta,
                'alerte':        delta >= seuil_jours,
                'etablissement': d.etablissement.nom if d.etablissement else 'N/A',
                'type_stage':    d.type_stage or '',
            })

        nb_alertes = sum(1 for r in rows if r['alerte'])
        filename = f"demandes_en_cours_{today}"

        if format == 'excel':
            wb, ws = make_excel_wb("Demandes en cours")
            start = excel_title_block(ws, f"Demandes en cours — {today.strftime('%d/%m/%Y')}",
                                      f"Seuil d'alerte : {seuil_jours} jours · {len(rows)} dossier(s) · {nb_alertes} alerte(s)",
                                      merge_cols=7)

            headers = ['N° Suivi', 'Demandeur', 'Statut', 'Date soumission', 'Jours écoulés', 'Alerte', 'Établissement']
            excel_header_row(ws, headers, row=start)
            for i, r in enumerate(rows, start + 1):
                vals = [r['tracking_id'], r['nom'], r['statut'], r['date_soumission'],
                        r['jours_ecoules'], 'RETARD' if r['alerte'] else '', r['etablissement']]
                excel_data_row(ws, vals, i, even=(i % 2 == 0))
                if r['alerte']:
                    ws.cell(i, 5).font = Font(color="EF4444", bold=True)
                    ws.cell(i, 6).font = Font(color="EF4444", bold=True)
                    ws.cell(i, 6).fill = PatternFill("solid", fgColor="FEE2E2")

            excel_autofit(ws)

            # Graphique répartition par statut
            if statut_counts:
                ws_chart = wb.create_sheet("Graphique Statuts")
                ws_chart['A1'] = 'Statut'
                ws_chart['B1'] = 'Nombre'
                for i, (s, c) in enumerate(statut_counts.items(), 2):
                    ws_chart.cell(i, 1, s)
                    ws_chart.cell(i, 2, c)
                excel_add_pie_chart(ws_chart, "Répartition par statut",
                                    {'min_col': 1, 'min_row': 1, 'max_row': len(statut_counts) + 1},
                                    {'min_col': 2, 'min_row': 1, 'max_row': len(statut_counts) + 1},
                                    anchor='D1')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer, filename

        # PDF
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, "Demandes en cours")
        styles = getSampleStyleSheet()
        elements = pdf_header("Demandes en cours",
                              f"Statut : {statut} · Seuil alerte : {seuil_jours} jours · {len(rows)} dossier(s) · {nb_alertes} alerte(s)", styles)

        # KPIs
        elements.extend(pdf_section_title("Indicateurs clés", styles))
        kpis = [
            ('Total dossiers en cours', len(rows)),
            ('Dossiers en alerte (≥ {} jours)'.format(seuil_jours), nb_alertes, CLR_DANGER if nb_alertes > 0 else None),
        ]
        for s, c in statut_counts.items():
            kpis.append((f'Statut : {s}', c))
        elements.append(pdf_kpi_table(kpis))
        elements.append(Spacer(1, 12))

        # Graphique répartition par statut
        if statut_counts:
            elements.extend(pdf_section_title("Répartition par statut", styles))
            elements.append(make_pie_chart(
                list(statut_counts.keys()), list(statut_counts.values()),
                "Répartition des demandes par statut"
            ))
            elements.append(Spacer(1, 12))

        # Tableau détaillé
        if rows:
            elements.extend(pdf_section_title("Détail des dossiers", styles))
            data = [['N° Suivi', 'Demandeur', 'Statut', 'Soumission', 'Jours', 'Alerte']]
            for r in rows:
                alerte_cell = Paragraph('<font color="red"><b>RETARD</b></font>', styles['Normal']) if r['alerte'] else ''
                data.append([r['tracking_id'], r['nom'], r['statut'],
                            r['date_soumission'], str(r['jours_ecoules']), alerte_cell])
            elements.append(pdf_table(data, [3*cm, 4*cm, 3*cm, 2.5*cm, 1.5*cm, 2.5*cm]))
        else:
            elements.append(Paragraph("Aucune demande en cours.", styles['Normal']))

        elements.append(pdf_footer_note(f"Rapport généré automatiquement — CEB Gestion des Stages — {today.strftime('%d/%m/%Y')}"))
        doc.build(elements)
        buffer.seek(0)
        return buffer, filename

    @staticmethod
    def generer_stages_actifs(params, format='pdf'):
        direction = params.get('direction', 'tous')
        today = date.today()

        qs = Stagiaire.objects.filter(statut='Actuel').select_related('etablissement')
        if direction != 'tous':
            qs = qs.filter(direction=direction)
        qs = qs.order_by('direction', 'service', 'nom')

        # Stats par direction
        par_dir = {}
        urgents = 0
        for s in qs:
            d = s.direction or 'N/A'
            par_dir[d] = par_dir.get(d, 0) + 1
            if s.date_fin and (s.date_fin - today).days < 15:
                urgents += 1

        filename = f"stages_actifs_{today}"

        if format == 'excel':
            wb, ws = make_excel_wb("Stages actifs")
            start = excel_title_block(ws, f"Stages actifs — {today.strftime('%d/%m/%Y')}",
                                      f"Direction : {direction} · {qs.count()} stagiaire(s) · {urgents} fin(s) imminente(s)",
                                      merge_cols=9)

            headers = ['Nom', 'Prénom', 'Direction', 'Service', 'Superviseur', 'Début', 'Fin', 'J. restants', 'Établissement']
            excel_header_row(ws, headers, row=start)
            for i, s in enumerate(qs, start + 1):
                jours = (s.date_fin - today).days if s.date_fin else 0
                excel_data_row(ws, [s.nom, s.prenom, s.direction, s.service, s.superviseur,
                    s.date_debut.strftime('%d/%m/%Y') if s.date_debut else '',
                    s.date_fin.strftime('%d/%m/%Y') if s.date_fin else '',
                    jours, s.etablissement.nom if s.etablissement else ''], i, even=(i % 2 == 0))
                if 0 <= jours < 15:
                    ws.cell(i, 8).font = Font(color="EF4444", bold=True)
                    ws.cell(i, 8).fill = PatternFill("solid", fgColor="FEE2E2")
            excel_autofit(ws)

            # Graphique par direction
            if par_dir:
                ws_chart = wb.create_sheet("Graphique Directions")
                ws_chart['A1'] = 'Direction'
                ws_chart['B1'] = 'Stagiaires'
                for i, (d, c) in enumerate(sorted(par_dir.items(), key=lambda x: -x[1]), 2):
                    ws_chart.cell(i, 1, d)
                    ws_chart.cell(i, 2, c)
                excel_add_bar_chart(ws_chart, "Répartition par direction",
                                    {'min_col': 1, 'min_row': 2, 'max_row': len(par_dir) + 1},
                                    {'min_col': 2, 'min_row': 1, 'max_row': len(par_dir) + 1},
                                    anchor='D1')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer, filename

        # PDF
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, "Stages actifs")
        styles = getSampleStyleSheet()
        elements = pdf_header("Stages actifs",
                              f"Direction : {direction} · {qs.count()} stagiaire(s) en cours", styles)

        # KPIs
        elements.extend(pdf_section_title("Indicateurs clés", styles))
        elements.append(pdf_kpi_table([
            ('Total stagiaires actifs', qs.count()),
            ('Fins imminentes (< 15 jours)', urgents, CLR_DANGER if urgents > 0 else None),
            ('Directions concernées', len(par_dir)),
        ]))
        elements.append(Spacer(1, 12))

        # Graphique par direction
        if par_dir:
            dirs_sorted = sorted(par_dir.items(), key=lambda x: -x[1])
            elements.extend(pdf_section_title("Répartition par direction", styles))
            elements.append(make_bar_chart(
                [d[0][:20] for d in dirs_sorted], [d[1] for d in dirs_sorted],
                "Stagiaires par direction", horizontal=True
            ))
            elements.append(Spacer(1, 12))

        # Tableau
        elements.extend(pdf_section_title("Liste détaillée", styles))
        data = [['Nom & Prénom', 'Direction / Service', 'Superviseur', 'Début', 'Fin', 'J. rest.']]
        for s in qs:
            jours = (s.date_fin - today).days if s.date_fin else 0
            jours_cell = Paragraph(f'<font color="red"><b>{jours}</b></font>', styles['Normal']) if jours < 15 else str(jours)
            data.append([f"{s.prenom} {s.nom}", f"{s.direction} / {s.service}", s.superviseur or '',
                s.date_debut.strftime('%d/%m/%Y') if s.date_debut else '',
                s.date_fin.strftime('%d/%m/%Y') if s.date_fin else '', jours_cell])
        elements.append(pdf_table(data, [4*cm, 4*cm, 3*cm, 2.5*cm, 2.5*cm, 1.5*cm]))
        elements.append(pdf_footer_note(f"Rapport généré automatiquement — CEB — {today.strftime('%d/%m/%Y')}"))
        doc.build(elements)
        buffer.seek(0)
        return buffer, filename

    @staticmethod
    def generer_fins_imminentes(params, format='pdf'):
        horizon = int(params.get('horizon_jours', 15))
        today   = date.today()
        limite  = today + timedelta(days=horizon)

        qs = Stagiaire.objects.filter(
            statut='Actuel', date_fin__gte=today, date_fin__lte=limite
        ).select_related('etablissement').order_by('date_fin')

        # Distribution par tranche de jours
        tranches = {'0-3 jours': 0, '4-7 jours': 0, '8-14 jours': 0, '15+ jours': 0}
        for s in qs:
            j = (s.date_fin - today).days
            if j <= 3:
                tranches['0-3 jours'] += 1
            elif j <= 7:
                tranches['4-7 jours'] += 1
            elif j <= 14:
                tranches['8-14 jours'] += 1
            else:
                tranches['15+ jours'] += 1

        filename = f"fins_imminentes_{today}"

        if format == 'excel':
            wb, ws = make_excel_wb("Fins imminentes")
            start = excel_title_block(ws,
                f"Fins de stage dans les {horizon} prochains jours",
                f"Date : {today.strftime('%d/%m/%Y')} · {qs.count()} stage(s) concerné(s)", merge_cols=7)

            headers = ['Stagiaire', 'Direction', 'Service', 'Date de fin', 'Jours restants', 'Email', 'Déjà renouvelé']
            excel_header_row(ws, headers, row=start, fill_color="EF4444")
            for i, s in enumerate(qs, start + 1):
                jours = (s.date_fin - today).days
                excel_data_row(ws, [f"{s.prenom} {s.nom}", s.direction, s.service,
                    s.date_fin.strftime('%d/%m/%Y'), jours, s.email,
                    'Oui' if getattr(s, 'a_ete_renouvele', False) else 'Non'], i, even=(i % 2 == 0))
                if jours <= 3:
                    for c in range(1, 8):
                        ws.cell(i, c).fill = PatternFill("solid", fgColor="FEE2E2")
            excel_autofit(ws)

            # Graphique tranches
            ws_chart = wb.create_sheet("Graphique Urgence")
            ws_chart['A1'] = 'Tranche'
            ws_chart['B1'] = 'Nombre'
            for i, (k, v) in enumerate(tranches.items(), 2):
                ws_chart.cell(i, 1, k)
                ws_chart.cell(i, 2, v)
            excel_add_bar_chart(ws_chart, "Distribution par urgence",
                                {'min_col': 1, 'min_row': 2, 'max_row': len(tranches) + 1},
                                {'min_col': 2, 'min_row': 1, 'max_row': len(tranches) + 1},
                                anchor='D1')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer, filename

        # PDF
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, "Fins de stage imminentes")
        styles = getSampleStyleSheet()
        elements = pdf_header("Fins de stage imminentes",
            f"Horizon : {horizon} jours · {qs.count()} stage(s) concerné(s)", styles)

        # KPIs
        elements.extend(pdf_section_title("Niveau d'urgence", styles, color=CLR_DANGER))
        elements.append(pdf_kpi_table([
            ('Total stages concernés', qs.count()),
            ('Très urgent (0-3 jours)', tranches['0-3 jours'], CLR_DANGER if tranches['0-3 jours'] > 0 else None),
            ('Urgent (4-7 jours)', tranches['4-7 jours'], CLR_WARNING if tranches['4-7 jours'] > 0 else None),
            ('À surveiller (8-14 jours)', tranches['8-14 jours']),
        ]))
        elements.append(Spacer(1, 12))

        # Graphique
        elements.extend(pdf_section_title("Distribution par urgence", styles))
        elements.append(make_bar_chart(
            list(tranches.keys()), list(tranches.values()),
            "Répartition par tranche de jours restants",
            color=CLR_DANGER
        ))
        elements.append(Spacer(1, 12))

        # Tableau
        elements.extend(pdf_section_title("Liste détaillée", styles))
        data = [['Stagiaire', 'Direction / Service', 'Date de fin', 'J. restants', 'Email']]
        for s in qs:
            jours = (s.date_fin - today).days
            data.append([f"{s.prenom} {s.nom}", f"{s.direction}/{s.service}",
                s.date_fin.strftime('%d/%m/%Y'),
                Paragraph(f'<font color="red"><b>{jours}</b></font>', styles['Normal']),
                s.email])
        elements.append(pdf_table(data, [4*cm, 4*cm, 2.5*cm, 2*cm, 5*cm], header_color=DANGER_COLOR))
        elements.append(pdf_footer_note(f"Rapport généré automatiquement — CEB — {today.strftime('%d/%m/%Y')}"))
        doc.build(elements)
        buffer.seek(0)
        return buffer, filename

    @staticmethod
    def generer_attestations_en_attente(params, format='pdf'):
        today = date.today()
        qs = DemandeAttestation.objects.filter(statut='en_attente').select_related('stagiaire').order_by('date_demande')

        # Distribution par ancienneté
        tranches = {'< 7 jours': 0, '7-14 jours': 0, '15-30 jours': 0, '> 30 jours': 0}
        for da in qs:
            j = (today - da.date_demande.date()).days if da.date_demande else 0
            if j < 7:
                tranches['< 7 jours'] += 1
            elif j <= 14:
                tranches['7-14 jours'] += 1
            elif j <= 30:
                tranches['15-30 jours'] += 1
            else:
                tranches['> 30 jours'] += 1

        filename = f"attestations_en_attente_{today}"

        if format == 'excel':
            wb, ws = make_excel_wb("Attestations en attente")
            start = excel_title_block(ws,
                f"Demandes d'attestation en attente",
                f"Date : {today.strftime('%d/%m/%Y')} · {qs.count()} demande(s)", merge_cols=6)

            excel_header_row(ws, ['Stagiaire', 'Direction', 'Service', 'Fin de stage', 'Date demande', "Jours d'attente"], row=start)
            for i, da in enumerate(qs, start + 1):
                s = da.stagiaire
                jours = (today - da.date_demande.date()).days if da.date_demande else 0
                excel_data_row(ws, [f"{s.prenom} {s.nom}", s.direction, s.service,
                    s.date_fin.strftime('%d/%m/%Y') if s.date_fin else '',
                    da.date_demande.strftime('%d/%m/%Y') if da.date_demande else '', jours], i, even=(i % 2 == 0))
                if jours > 14:
                    ws.cell(i, 6).font = Font(color="EF4444", bold=True)
            excel_autofit(ws)

            # Graphique
            ws_chart = wb.create_sheet("Graphique Attente")
            ws_chart['A1'] = 'Ancienneté'
            ws_chart['B1'] = 'Nombre'
            for i, (k, v) in enumerate(tranches.items(), 2):
                ws_chart.cell(i, 1, k)
                ws_chart.cell(i, 2, v)
            excel_add_pie_chart(ws_chart, "Ancienneté des demandes",
                                {'min_col': 1, 'min_row': 1, 'max_row': len(tranches) + 1},
                                {'min_col': 2, 'min_row': 1, 'max_row': len(tranches) + 1},
                                anchor='D1')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer, filename

        # PDF
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, "Attestations en attente")
        styles = getSampleStyleSheet()
        elements = pdf_header("Attestations en attente", f"{qs.count()} demande(s) non traitée(s)", styles)

        elements.extend(pdf_section_title("Indicateurs", styles))
        elements.append(pdf_kpi_table([
            ('Total en attente', qs.count()),
            ('Attente > 14 jours', tranches['15-30 jours'] + tranches['> 30 jours'],
             CLR_DANGER if (tranches['15-30 jours'] + tranches['> 30 jours']) > 0 else None),
        ]))
        elements.append(Spacer(1, 12))

        # Graphique
        elements.extend(pdf_section_title("Ancienneté des demandes", styles))
        elements.append(make_pie_chart(
            list(tranches.keys()), list(tranches.values()),
            "Répartition par ancienneté"
        ))
        elements.append(Spacer(1, 12))

        # Tableau
        elements.extend(pdf_section_title("Liste détaillée", styles))
        data = [['Stagiaire', 'Direction', 'Date demande', "Jours d'attente"]]
        for da in qs:
            s = da.stagiaire
            jours = (today - da.date_demande.date()).days if da.date_demande else 0
            data.append([f"{s.prenom} {s.nom}", f"{s.direction}/{s.service}",
                da.date_demande.strftime('%d/%m/%Y') if da.date_demande else '', str(jours)])
        elements.append(pdf_table(data, [5*cm, 4*cm, 3.5*cm, 3*cm]))
        elements.append(pdf_footer_note(f"Rapport généré automatiquement — CEB — {today.strftime('%d/%m/%Y')}"))
        doc.build(elements)
        buffer.seek(0)
        return buffer, filename

    # ─────────────────────────────────────────────────────────────────────────
    # PERFORMANCE
    # ─────────────────────────────────────────────────────────────────────────

    @staticmethod
    def generer_synthese_mensuelle(params, format='pdf'):
        annee = int(params.get('annee', date.today().year))
        mois  = int(params.get('mois',  date.today().month))
        today = date.today()

        qs_mois = Demande.objects.filter(date_soumission__year=annee, date_soumission__month=mois)
        total       = qs_mois.count()
        acceptees   = qs_mois.filter(statut_stage='Acceptée').count()
        refusees    = qs_mois.filter(statut_stage='Refusée').count()
        en_cours    = qs_mois.exclude(statut_stage__in=['Acceptée', 'Refusée']).count()
        taux_acc    = round(acceptees / total * 100, 1) if total else 0

        # Mois précédent
        if mois == 1:
            mois_prec, annee_prec = 12, annee - 1
        else:
            mois_prec, annee_prec = mois - 1, annee
        qs_prec    = Demande.objects.filter(date_soumission__year=annee_prec, date_soumission__month=mois_prec)
        total_prec = qs_prec.count()
        acc_prec   = qs_prec.filter(statut_stage='Acceptée').count()
        ref_prec   = qs_prec.filter(statut_stage='Refusée').count()
        ec_prec    = qs_prec.exclude(statut_stage__in=['Acceptée', 'Refusée']).count()
        taux_prec  = round(acc_prec / total_prec * 100, 1) if total_prec else 0

        nom_mois = NOMS_MOIS[mois]
        nom_mois_prec = NOMS_MOIS[mois_prec]
        filename = f"synthese_{nom_mois.lower()}_{annee}"

        kpis = [
            ('Total demandes', total, total_prec),
            ('Demandes acceptées', acceptees, acc_prec),
            ('Demandes refusées', refusees, ref_prec),
            ('En cours', en_cours, ec_prec),
            ("Taux d'acceptation", f"{taux_acc}%", f"{taux_prec}%"),
        ]

        # Données pour graphiques comparatifs
        labels_comp = [nom_mois_prec, nom_mois]
        vals_total = [total_prec, total]
        vals_acc   = [acc_prec, acceptees]
        vals_ref   = [ref_prec, refusees]

        if format == 'excel':
            wb, ws = make_excel_wb(f"Synthèse {nom_mois}")
            start = excel_title_block(ws, f"Synthèse mensuelle — {nom_mois} {annee}",
                                      f"Comparaison avec {nom_mois_prec} {annee_prec}", merge_cols=4)

            excel_header_row(ws, ['Indicateur', f'{nom_mois} {annee}', 'Mois précédent', 'Évolution'], row=start)
            for i, (label, val, prec) in enumerate(kpis, start + 1):
                try:
                    v1 = int(str(val).replace('%', ''))
                    v2 = int(str(prec).replace('%', ''))
                    evol = f"+{v1 - v2}" if v1 >= v2 else str(v1 - v2)
                except (ValueError, TypeError):
                    evol = '—'
                excel_data_row(ws, [label, val, prec, evol], i, even=(i % 2 == 0))
            excel_autofit(ws)

            # Graphique comparatif
            ws_chart = wb.create_sheet("Graphique Comparaison")
            ws_chart['A1'] = 'Catégorie'
            ws_chart['B1'] = nom_mois_prec
            ws_chart['C1'] = nom_mois
            categories = ['Total', 'Acceptées', 'Refusées', 'En cours']
            for i, (cat, vp, vc) in enumerate(zip(categories, [total_prec, acc_prec, ref_prec, ec_prec],
                                                               [total, acceptees, refusees, en_cours]), 2):
                ws_chart.cell(i, 1, cat)
                ws_chart.cell(i, 2, vp)
                ws_chart.cell(i, 3, vc)

            chart = BarChart()
            chart.type = "col"
            chart.title = "Comparaison mensuelle"
            chart.width = 18
            chart.height = 12
            d1 = Reference(ws_chart, min_col=2, min_row=1, max_row=5)
            d2 = Reference(ws_chart, min_col=3, min_row=1, max_row=5)
            cats = Reference(ws_chart, min_col=1, min_row=2, max_row=5)
            chart.add_data(d1, titles_from_data=True)
            chart.add_data(d2, titles_from_data=True)
            chart.set_categories(cats)
            chart.series[0].graphicalProperties.solidFill = '94A3B8'
            chart.series[1].graphicalProperties.solidFill = '6366F1'
            ws_chart.add_chart(chart, 'E1')

            # Pie chart statuts mois courant
            ws_pie = wb.create_sheet("Graphique Statuts")
            ws_pie['A1'] = 'Statut'
            ws_pie['B1'] = 'Nombre'
            for i, (s, v) in enumerate([('Acceptées', acceptees), ('Refusées', refusees), ('En cours', en_cours)], 2):
                ws_pie.cell(i, 1, s)
                ws_pie.cell(i, 2, v)
            excel_add_pie_chart(ws_pie, f"Répartition {nom_mois} {annee}",
                                {'min_col': 1, 'min_row': 1, 'max_row': 4},
                                {'min_col': 2, 'min_row': 1, 'max_row': 4},
                                anchor='D1')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer, filename

        # PDF
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, f"Synthèse mensuelle — {nom_mois} {annee}")
        styles = getSampleStyleSheet()
        elements = pdf_header(f"Synthèse mensuelle — {nom_mois} {annee}",
                              f"Comparaison avec {nom_mois_prec} {annee_prec}", styles)

        # KPIs avec évolution
        elements.extend(pdf_section_title("Indicateurs clés", styles))
        data = [['Indicateur', f'{nom_mois} {annee}', 'Mois précédent', 'Évolution']]
        for label, val, prec in kpis:
            try:
                diff = int(str(val).replace('%', '')) - int(str(prec).replace('%', ''))
                evol = Paragraph(f'<font color="{"#10b981" if diff >= 0 else "#ef4444"}"><b>{"+" if diff >= 0 else ""}{diff}</b></font>', styles['Normal'])
            except (ValueError, TypeError):
                evol = '—'
            data.append([label, str(val), str(prec), evol])
        elements.append(pdf_table(data, [5.5*cm, 3*cm, 3*cm, 3*cm]))
        elements.append(Spacer(1, 16))

        # Graphique comparatif barres
        elements.extend(pdf_section_title("Comparaison visuelle", styles))
        _setup_chart_style()
        fig, ax = plt.subplots(figsize=(5.5, 3))
        x_pos = range(4)
        categories = ['Total', 'Acceptées', 'Refusées', 'En cours']
        w = 0.35
        bars1 = ax.bar([p - w/2 for p in x_pos], [total_prec, acc_prec, ref_prec, ec_prec],
                        w, label=nom_mois_prec, color='#94a3b8')
        bars2 = ax.bar([p + w/2 for p in x_pos], [total, acceptees, refusees, en_cours],
                        w, label=nom_mois, color=CLR_ACCENT)
        ax.set_xticks(x_pos)
        ax.set_xticklabels(categories)
        ax.set_title("Comparaison mensuelle")
        ax.legend(frameon=False, fontsize=8)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
        for bars in [bars1, bars2]:
            for bar in bars:
                h = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2, h + 0.1, str(int(h)),
                        ha='center', fontsize=7, fontweight='bold')
        plt.tight_layout()
        elements.append(_chart_to_image(fig, 16*cm, 8*cm))
        elements.append(Spacer(1, 12))

        # Pie chart
        elements.extend(pdf_section_title(f"Répartition des statuts — {nom_mois}", styles))
        elements.append(make_pie_chart(
            ['Acceptées', 'Refusées', 'En cours'],
            [acceptees, refusees, en_cours],
            f"Statuts des demandes — {nom_mois} {annee}"
        ))

        elements.append(pdf_footer_note(f"Rapport généré automatiquement — CEB — {today.strftime('%d/%m/%Y')}"))
        doc.build(elements)
        buffer.seek(0)
        return buffer, filename

    @staticmethod
    def generer_rapport_annuel(params, format='pdf'):
        annee = int(params.get('annee', date.today().year))
        today = date.today()
        qs = Demande.objects.filter(date_soumission__year=annee)

        total     = qs.count()
        acceptees = qs.filter(statut_stage='Acceptée').count()
        refusees  = qs.filter(statut_stage='Refusée').count()
        en_cours  = qs.exclude(statut_stage__in=['Acceptée', 'Refusée']).count()
        taux      = round(acceptees / total * 100, 1) if total else 0

        # Par mois
        par_mois = [qs.filter(date_soumission__month=m + 1).count() for m in range(12)]

        # Par direction (top 8)
        par_direction = list(
            Stagiaire.objects.filter(date_debut__year=annee)
            .values('direction').annotate(total=Count('id')).order_by('-total')[:8]
        )

        # Par statut
        statut_data = {'Acceptées': acceptees, 'Refusées': refusees, 'En cours': en_cours}

        filename = f"rapport_annuel_{annee}"

        if format == 'excel':
            wb, ws = make_excel_wb(f"Rapport {annee}")
            start = excel_title_block(ws, f"Rapport annuel {annee}",
                                      f"Total : {total} demande(s) · Taux d'acceptation : {taux}%", merge_cols=3)

            # KPIs
            ws.cell(start, 1, "Récapitulatif").font = Font(bold=True, size=11, color="6366F1")
            excel_header_row(ws, ['Indicateur', 'Valeur'], row=start + 1)
            kpi_data = [('Total demandes', total), ('Acceptées', acceptees), ('Refusées', refusees),
                        ('En cours', en_cours), ("Taux d'acceptation", f"{taux}%")]
            for i, (k, v) in enumerate(kpi_data, start + 2):
                excel_data_row(ws, [k, v], i, even=(i % 2 == 0))
            excel_autofit(ws)

            # Feuille évolution mensuelle
            ws_mois = wb.create_sheet("Évolution Mensuelle")
            ws_mois['A1'] = 'Mois'
            ws_mois['B1'] = 'Demandes'
            excel_header_row(ws_mois, ['Mois', 'Demandes'], row=1)
            for i, (m, v) in enumerate(zip(NOMS_MOIS_COURT, par_mois), 2):
                ws_mois.cell(i, 1, m)
                ws_mois.cell(i, 2, v)
            excel_autofit(ws_mois)
            excel_add_bar_chart(ws_mois, f"Évolution mensuelle {annee}",
                                {'min_col': 1, 'min_row': 2, 'max_row': 13},
                                {'min_col': 2, 'min_row': 1, 'max_row': 13},
                                anchor='D1')

            # Feuille répartition statuts
            ws_stat = wb.create_sheet("Statuts")
            ws_stat['A1'] = 'Statut'
            ws_stat['B1'] = 'Nombre'
            for i, (s, v) in enumerate(statut_data.items(), 2):
                ws_stat.cell(i, 1, s)
                ws_stat.cell(i, 2, v)
            excel_add_pie_chart(ws_stat, f"Répartition par statut {annee}",
                                {'min_col': 1, 'min_row': 1, 'max_row': 4},
                                {'min_col': 2, 'min_row': 1, 'max_row': 4},
                                anchor='D1')

            # Feuille directions
            if par_direction:
                ws_dir = wb.create_sheet("Directions")
                ws_dir['A1'] = 'Direction'
                ws_dir['B1'] = 'Stagiaires'
                excel_header_row(ws_dir, ['Direction', 'Stagiaires'], row=1)
                for i, d in enumerate(par_direction, 2):
                    ws_dir.cell(i, 1, d['direction'] or 'N/A')
                    ws_dir.cell(i, 2, d['total'])
                excel_autofit(ws_dir)
                excel_add_bar_chart(ws_dir, "Répartition par direction",
                                    {'min_col': 1, 'min_row': 2, 'max_row': len(par_direction) + 1},
                                    {'min_col': 2, 'min_row': 1, 'max_row': len(par_direction) + 1},
                                    anchor='D1')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer, filename

        # PDF
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, f"Rapport annuel {annee}")
        styles = getSampleStyleSheet()
        elements = pdf_header(f"Rapport annuel {annee}",
                              f"{total} demande(s) traitée(s) — Taux d'acceptation : {taux}%", styles)

        # KPIs
        elements.extend(pdf_section_title("Récapitulatif général", styles))
        elements.append(pdf_kpi_table([
            ('Total demandes', total),
            ('Acceptées', f"{acceptees} ({taux}%)", CLR_SUCCESS),
            ('Refusées', refusees, CLR_DANGER),
            ('En cours de traitement', en_cours, CLR_WARNING),
        ]))
        elements.append(Spacer(1, 16))

        # Pie chart statuts
        elements.extend(pdf_section_title("Répartition par statut", styles))
        elements.append(make_pie_chart(
            list(statut_data.keys()), list(statut_data.values()),
            f"Statuts des demandes — {annee}"
        ))
        elements.append(Spacer(1, 16))

        # Évolution mensuelle (line chart)
        elements.extend(pdf_section_title("Évolution mensuelle", styles))
        elements.append(make_line_chart(
            NOMS_MOIS_COURT, [('Demandes', par_mois)],
            f"Évolution mensuelle des demandes — {annee}"
        ))
        elements.append(Spacer(1, 16))

        # Bar chart directions
        if par_direction:
            elements.extend(pdf_section_title("Répartition par direction", styles))
            dir_labels = [(d['direction'] or 'N/A')[:20] for d in par_direction]
            dir_values = [d['total'] for d in par_direction]
            elements.append(make_bar_chart(
                dir_labels, dir_values,
                f"Top {len(par_direction)} directions — {annee}",
                horizontal=True
            ))

        elements.append(pdf_footer_note(f"Rapport annuel généré automatiquement — CEB Gestion des Stages — {today.strftime('%d/%m/%Y')}"))
        doc.build(elements)
        buffer.seek(0)
        return buffer, filename

    @staticmethod
    def generer_performance_traitement(params, format='pdf'):
        annee = int(params.get('annee', date.today().year))
        mois  = params.get('mois', 'tous')
        today = date.today()

        qs = Demande.objects.filter(date_soumission__year=annee)
        if mois != 'tous':
            qs = qs.filter(date_soumission__month=int(mois))

        # Par jour de la semaine
        from django.db.models.functions import ExtractWeekDay
        hebdo = (qs.annotate(jour=ExtractWeekDay('date_soumission'))
                   .values('jour').annotate(total=Count('id')).order_by('jour'))
        hebdo_data = {h['jour']: h['total'] for h in hebdo}

        # Ordre: Lun-Dim
        jours_order = [2, 3, 4, 5, 6, 7, 1]
        jours_names = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
        jours_vals  = [hebdo_data.get(j, 0) for j in jours_order]

        total_semaine = sum(jours_vals)
        jour_max_idx = jours_vals.index(max(jours_vals)) if jours_vals else 0
        jour_max = jours_names[jour_max_idx] if jours_vals else 'N/A'

        # Par mois
        par_mois = [qs.filter(date_soumission__month=m + 1).count() for m in range(12)]

        filename = f"performance_traitement_{annee}"

        if format == 'excel':
            wb, ws = make_excel_wb("Performance")
            start = excel_title_block(ws,
                f"Performance de traitement — {annee}",
                f"Mois : {'Tous' if mois == 'tous' else NOMS_MOIS[int(mois)]} · Pic : {jour_max}",
                merge_cols=3)

            excel_header_row(ws, ['Jour de la semaine', 'Demandes reçues', '% du total'], row=start, fill_color="F97316")
            for i, (j, v) in enumerate(zip(jours_names, jours_vals), start + 1):
                pct = round(v / total_semaine * 100, 1) if total_semaine else 0
                excel_data_row(ws, [j, v, f"{pct}%"], i, even=(i % 2 == 0))
            excel_autofit(ws)

            # Graphique hebdomadaire
            ws_chart = wb.create_sheet("Graphique Hebdo")
            ws_chart['A1'] = 'Jour'
            ws_chart['B1'] = 'Demandes'
            for i, (j, v) in enumerate(zip(jours_names, jours_vals), 2):
                ws_chart.cell(i, 1, j)
                ws_chart.cell(i, 2, v)
            excel_add_bar_chart(ws_chart, "Performance hebdomadaire",
                                {'min_col': 1, 'min_row': 2, 'max_row': 8},
                                {'min_col': 2, 'min_row': 1, 'max_row': 8},
                                anchor='D1')

            # Graphique mensuel
            if mois == 'tous':
                ws_mois = wb.create_sheet("Graphique Mensuel")
                ws_mois['A1'] = 'Mois'
                ws_mois['B1'] = 'Demandes'
                for i, (m, v) in enumerate(zip(NOMS_MOIS_COURT, par_mois), 2):
                    ws_mois.cell(i, 1, m)
                    ws_mois.cell(i, 2, v)
                excel_add_bar_chart(ws_mois, f"Évolution mensuelle {annee}",
                                    {'min_col': 1, 'min_row': 2, 'max_row': 13},
                                    {'min_col': 2, 'min_row': 1, 'max_row': 13},
                                    anchor='D1')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer, filename

        # PDF
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, "Performance de traitement")
        styles = getSampleStyleSheet()
        elements = pdf_header("Performance de traitement",
            f"Année {annee}" + (f" — {NOMS_MOIS[int(mois)]}" if mois != 'tous' else " — Tous les mois"), styles)

        # KPIs
        elements.extend(pdf_section_title("Indicateurs clés", styles))
        elements.append(pdf_kpi_table([
            ('Total demandes', total_semaine),
            ('Jour le plus actif', jour_max, CLR_ACCENT),
            ('Moyenne par jour', round(total_semaine / 7, 1) if total_semaine else 0),
        ]))
        elements.append(Spacer(1, 14))

        # Graphique hebdomadaire
        elements.extend(pdf_section_title("Répartition hebdomadaire", styles))
        elements.append(make_bar_chart(
            jours_names, jours_vals,
            "Demandes par jour de la semaine",
            color=CLR_WARNING
        ))
        elements.append(Spacer(1, 12))

        # Tableau hebdo
        data = [['Jour', 'Demandes', '% du total']]
        for j, v in zip(jours_names, jours_vals):
            pct = round(v / total_semaine * 100, 1) if total_semaine else 0
            data.append([j, str(v), f"{pct}%"])
        elements.append(pdf_table(data, [6*cm, 5*cm, 5*cm]))
        elements.append(Spacer(1, 14))

        # Graphique mensuel
        if mois == 'tous':
            elements.extend(pdf_section_title("Évolution mensuelle", styles))
            elements.append(make_line_chart(
                NOMS_MOIS_COURT, [('Demandes', par_mois)],
                f"Tendance mensuelle — {annee}"
            ))

        elements.append(pdf_footer_note(f"Rapport généré automatiquement — CEB — {today.strftime('%d/%m/%Y')}"))
        doc.build(elements)
        buffer.seek(0)
        return buffer, filename

    # ─────────────────────────────────────────────────────────────────────────
    # RH
    # ─────────────────────────────────────────────────────────────────────────

    @staticmethod
    def generer_repartition_directions(params, format='pdf'):
        annee  = int(params.get('annee', date.today().year))
        statut = params.get('statut', 'tous')
        today  = date.today()

        qs = Stagiaire.objects.filter(date_debut__year=annee)
        if statut != 'tous':
            qs = qs.filter(statut=statut)

        par_dir = list(qs.values('direction', 'service').annotate(total=Count('id')).order_by('direction', 'service'))
        # Agrégation par direction seule pour graphiques
        dir_agg = {}
        for row in par_dir:
            d = row['direction'] or 'N/A'
            dir_agg[d] = dir_agg.get(d, 0) + row['total']
        total_stagiaires = sum(dir_agg.values())

        filename = f"repartition_directions_{annee}"

        if format == 'excel':
            wb, ws = make_excel_wb("Directions")
            start = excel_title_block(ws, f"Répartition par direction — {annee}",
                                      f"Statut : {statut} · {total_stagiaires} stagiaire(s)", merge_cols=4)

            excel_header_row(ws, ['Direction', 'Service', 'Nombre de stagiaires', '% du total'], row=start)
            for i, row in enumerate(par_dir, start + 1):
                pct = round(row['total'] / total_stagiaires * 100, 1) if total_stagiaires else 0
                excel_data_row(ws, [row['direction'] or 'N/A', row['service'] or 'N/A',
                                    row['total'], f"{pct}%"], i, even=(i % 2 == 0))
            excel_autofit(ws)

            # Graphique
            ws_chart = wb.create_sheet("Graphique")
            ws_chart['A1'] = 'Direction'
            ws_chart['B1'] = 'Stagiaires'
            for i, (d, c) in enumerate(sorted(dir_agg.items(), key=lambda x: -x[1]), 2):
                ws_chart.cell(i, 1, d)
                ws_chart.cell(i, 2, c)
            excel_add_pie_chart(ws_chart, f"Répartition par direction — {annee}",
                                {'min_col': 1, 'min_row': 1, 'max_row': len(dir_agg) + 1},
                                {'min_col': 2, 'min_row': 1, 'max_row': len(dir_agg) + 1},
                                anchor='D1')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer, filename

        # PDF
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, "Répartition par direction")
        styles = getSampleStyleSheet()
        elements = pdf_header("Répartition par direction",
                              f"Année {annee} · Statut : {statut} · {total_stagiaires} stagiaire(s)", styles)

        # KPIs
        elements.extend(pdf_section_title("Synthèse", styles))
        elements.append(pdf_kpi_table([
            ('Total stagiaires', total_stagiaires),
            ('Nombre de directions', len(dir_agg)),
            ('Direction principale', max(dir_agg, key=dir_agg.get) if dir_agg else 'N/A', CLR_ACCENT),
        ]))
        elements.append(Spacer(1, 14))

        # Pie chart
        if dir_agg:
            elements.extend(pdf_section_title("Graphique de répartition", styles))
            sorted_dirs = sorted(dir_agg.items(), key=lambda x: -x[1])
            elements.append(make_pie_chart(
                [d[0][:25] for d in sorted_dirs], [d[1] for d in sorted_dirs],
                f"Répartition par direction — {annee}", width=14*cm, height=10*cm
            ))
            elements.append(Spacer(1, 14))

        # Tableau détaillé
        elements.extend(pdf_section_title("Détail par direction et service", styles))
        data = [['Direction', 'Service', 'Stagiaires', '% du total']]
        for row in par_dir:
            pct = round(row['total'] / total_stagiaires * 100, 1) if total_stagiaires else 0
            data.append([row['direction'] or 'N/A', row['service'] or 'N/A', str(row['total']), f"{pct}%"])
        elements.append(pdf_table(data, [4*cm, 5*cm, 3*cm, 3*cm]))
        elements.append(pdf_footer_note(f"Rapport généré automatiquement — CEB — {today.strftime('%d/%m/%Y')}"))
        doc.build(elements)
        buffer.seek(0)
        return buffer, filename

    @staticmethod
    def generer_profil_stagiaires(params, format='pdf'):
        annee = int(params.get('annee', date.today().year))
        qs = Stagiaire.objects.filter(date_debut__year=annee)
        today = date.today()
        total = qs.count()

        par_genre      = list(qs.values('genre').annotate(total=Count('id')))
        par_niveau     = list(qs.values('niveau_etude').annotate(total=Count('id')).order_by('-total'))
        par_specialite = list(qs.values('specialite').annotate(total=Count('id')).order_by('-total')[:10])
        par_pays       = list(qs.values('pays_residence').annotate(total=Count('id')).order_by('-total')[:10])

        filename = f"profil_stagiaires_{annee}"

        if format == 'excel':
            wb, ws = make_excel_wb("Profil")
            start = excel_title_block(ws, f"Profil démographique des stagiaires — {annee}",
                                      f"{total} stagiaire(s) analysé(s)", merge_cols=3)

            sections = [
                ("Par genre", par_genre, 'genre'),
                ("Par niveau d'études", par_niveau, 'niveau_etude'),
                ("Top 10 spécialités", par_specialite, 'specialite'),
                ("Top 10 pays", par_pays, 'pays_residence'),
            ]
            row = start
            for titre, data_qs, key in sections:
                ws.cell(row, 1, titre).font = Font(bold=True, size=10, color="6366F1")
                row += 1
                excel_header_row(ws, [key.replace('_', ' ').capitalize(), 'Total', '%'], row=row)
                row += 1
                for i, item in enumerate(data_qs):
                    pct = round(item['total'] / total * 100, 1) if total else 0
                    excel_data_row(ws, [item[key] or 'N/A', item['total'], f"{pct}%"],
                                  row, even=(i % 2 == 0))
                    row += 1
                row += 1
            excel_autofit(ws)

            # Graphique genre
            if par_genre:
                ws_g = wb.create_sheet("Graphique Genre")
                ws_g['A1'] = 'Genre'
                ws_g['B1'] = 'Nombre'
                for i, item in enumerate(par_genre, 2):
                    ws_g.cell(i, 1, item['genre'] or 'N/A')
                    ws_g.cell(i, 2, item['total'])
                excel_add_pie_chart(ws_g, "Répartition par genre",
                                    {'min_col': 1, 'min_row': 1, 'max_row': len(par_genre) + 1},
                                    {'min_col': 2, 'min_row': 1, 'max_row': len(par_genre) + 1},
                                    anchor='D1')

            # Graphique niveau
            if par_niveau:
                ws_n = wb.create_sheet("Graphique Niveau")
                ws_n['A1'] = 'Niveau'
                ws_n['B1'] = 'Nombre'
                for i, item in enumerate(par_niveau, 2):
                    ws_n.cell(i, 1, item['niveau_etude'] or 'N/A')
                    ws_n.cell(i, 2, item['total'])
                excel_add_bar_chart(ws_n, "Répartition par niveau d'études",
                                    {'min_col': 1, 'min_row': 2, 'max_row': len(par_niveau) + 1},
                                    {'min_col': 2, 'min_row': 1, 'max_row': len(par_niveau) + 1},
                                    anchor='D1')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer, filename

        # PDF
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, "Profil démographique")
        styles = getSampleStyleSheet()
        elements = pdf_header("Profil démographique des stagiaires",
                              f"Année {annee} · {total} stagiaire(s)", styles)

        # KPIs
        elements.extend(pdf_section_title("Vue d'ensemble", styles))
        genre_str = ' / '.join([f"{g['genre'] or 'N/A'}: {g['total']}" for g in par_genre])
        elements.append(pdf_kpi_table([
            ('Total stagiaires', total),
            ('Répartition genre', genre_str),
            ('Niveaux distincts', len(par_niveau)),
            ('Spécialités distinctes', len(par_specialite)),
        ]))
        elements.append(Spacer(1, 14))

        # Graphique genre
        if par_genre:
            elements.extend(pdf_section_title("Répartition par genre", styles))
            elements.append(make_pie_chart(
                [g['genre'] or 'N/A' for g in par_genre],
                [g['total'] for g in par_genre],
                "Genre"
            ))
            elements.append(Spacer(1, 14))

        # Graphique niveau
        if par_niveau:
            elements.extend(pdf_section_title("Répartition par niveau d'études", styles))
            elements.append(make_bar_chart(
                [(n['niveau_etude'] or 'N/A')[:15] for n in par_niveau],
                [n['total'] for n in par_niveau],
                "Niveau d'études", horizontal=True
            ))
            elements.append(Spacer(1, 14))

        # Graphique spécialités
        if par_specialite:
            elements.extend(pdf_section_title("Top 10 spécialités", styles))
            elements.append(make_bar_chart(
                [(s['specialite'] or 'N/A')[:20] for s in par_specialite],
                [s['total'] for s in par_specialite],
                "Spécialités les plus fréquentes", horizontal=True
            ))
            elements.append(Spacer(1, 14))

        # Tableau pays
        if par_pays:
            elements.extend(pdf_section_title("Top 10 pays de résidence", styles))
            data = [['Pays', 'Stagiaires', '%']]
            for item in par_pays:
                pct = round(item['total'] / total * 100, 1) if total else 0
                data.append([item['pays_residence'] or 'N/A', str(item['total']), f"{pct}%"])
            elements.append(pdf_table(data, [8*cm, 4*cm, 4*cm]))

        elements.append(pdf_footer_note(f"Rapport généré automatiquement — CEB — {today.strftime('%d/%m/%Y')}"))
        doc.build(elements)
        buffer.seek(0)
        return buffer, filename

    @staticmethod
    def generer_etablissements_partenaires(params, format='pdf'):
        annee = int(params.get('annee', date.today().year))
        top   = params.get('top', '10')
        today = date.today()

        qs = (Demande.objects
              .filter(date_soumission__year=annee, etablissement__isnull=False)
              .values('etablissement__nom', 'etablissement__email')
              .annotate(
                  total=Count('id'),
                  acceptees=Count('id', filter=Q(statut_stage='Acceptée')),
              ).order_by('-total'))
        if top != 'tous':
            qs = qs[:int(top)]

        etab_list = list(qs)
        filename = f"etablissements_partenaires_{annee}"

        if format == 'excel':
            wb, ws = make_excel_wb("Établissements")
            start = excel_title_block(ws, f"Établissements partenaires — {annee}",
                                      f"Top {top} · {len(etab_list)} établissement(s)", merge_cols=5)

            excel_header_row(ws, ['Établissement', 'Email', 'Demandes', 'Acceptées', "Taux d'acc."], row=start)
            for i, e in enumerate(etab_list, start + 1):
                taux = round(e['acceptees'] / e['total'] * 100, 1) if e['total'] else 0
                excel_data_row(ws, [e['etablissement__nom'], e['etablissement__email'] or '',
                                    e['total'], e['acceptees'], f"{taux}%"], i, even=(i % 2 == 0))
            excel_autofit(ws)

            # Graphique
            if etab_list:
                ws_chart = wb.create_sheet("Graphique")
                ws_chart['A1'] = 'Établissement'
                ws_chart['B1'] = 'Demandes'
                ws_chart['C1'] = 'Acceptées'
                for i, e in enumerate(etab_list[:10], 2):
                    ws_chart.cell(i, 1, (e['etablissement__nom'] or 'N/A')[:30])
                    ws_chart.cell(i, 2, e['total'])
                    ws_chart.cell(i, 3, e['acceptees'])
                n = min(len(etab_list), 10)

                chart = BarChart()
                chart.type = "col"
                chart.title = f"Top établissements — {annee}"
                chart.width = 20
                chart.height = 12
                d1 = Reference(ws_chart, min_col=2, min_row=1, max_row=n + 1)
                d2 = Reference(ws_chart, min_col=3, min_row=1, max_row=n + 1)
                cats = Reference(ws_chart, min_col=1, min_row=2, max_row=n + 1)
                chart.add_data(d1, titles_from_data=True)
                chart.add_data(d2, titles_from_data=True)
                chart.set_categories(cats)
                chart.series[0].graphicalProperties.solidFill = '6366F1'
                chart.series[1].graphicalProperties.solidFill = '10B981'
                ws_chart.add_chart(chart, 'E1')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer, filename

        # PDF
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, "Établissements partenaires")
        styles = getSampleStyleSheet()
        elements = pdf_header("Établissements partenaires",
                              f"Année {annee} · Top {top}", styles)

        # KPIs
        if etab_list:
            total_dem = sum(e['total'] for e in etab_list)
            total_acc = sum(e['acceptees'] for e in etab_list)
            elements.extend(pdf_section_title("Vue d'ensemble", styles))
            elements.append(pdf_kpi_table([
                ('Établissements listés', len(etab_list)),
                ('Total demandes', total_dem),
                ('Total acceptées', total_acc, CLR_SUCCESS),
                ('Taux moyen', f"{round(total_acc / total_dem * 100, 1)}%" if total_dem else '0%'),
            ]))
            elements.append(Spacer(1, 14))

            # Graphique
            elements.extend(pdf_section_title("Classement visuel", styles))
            top_10 = etab_list[:10]
            _setup_chart_style()
            fig, ax = plt.subplots(figsize=(6, 4))
            names = [(e['etablissement__nom'] or 'N/A')[:25] for e in top_10]
            x_pos = range(len(names))
            w = 0.35
            ax.barh([p - w/2 for p in x_pos], [e['total'] for e in top_10],
                    w, label='Demandes', color=CLR_ACCENT)
            ax.barh([p + w/2 for p in x_pos], [e['acceptees'] for e in top_10],
                    w, label='Acceptées', color=CLR_SUCCESS)
            ax.set_yticks(x_pos)
            ax.set_yticklabels(names, fontsize=7)
            ax.invert_yaxis()
            ax.set_title("Demandes vs Acceptées par établissement")
            ax.legend(frameon=False, fontsize=8)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            plt.tight_layout()
            elements.append(_chart_to_image(fig, 16*cm, 10*cm))
            elements.append(Spacer(1, 14))

        # Tableau
        elements.extend(pdf_section_title("Tableau détaillé", styles))
        data = [['Établissement', 'Demandes', 'Acceptées', "Taux d'acc."]]
        for e in etab_list:
            taux = round(e['acceptees'] / e['total'] * 100, 1) if e['total'] else 0
            data.append([e['etablissement__nom'], str(e['total']), str(e['acceptees']), f"{taux}%"])
        elements.append(pdf_table(data, [7*cm, 2.5*cm, 2.5*cm, 3*cm]))
        elements.append(pdf_footer_note(f"Rapport généré automatiquement — CEB — {today.strftime('%d/%m/%Y')}"))
        doc.build(elements)
        buffer.seek(0)
        return buffer, filename

    @staticmethod
    def generer_renouvellements(params, format='pdf'):
        annee = int(params.get('annee', date.today().year))
        today = date.today()

        qs = Stagiaire.objects.filter(
            est_renouvellement=True, date_debut__year=annee
        ).select_related('stage_precedent', 'etablissement').order_by('-date_debut')

        # Stats par direction
        par_dir = {}
        par_mois = [0] * 12
        for s in qs:
            d = s.direction or 'N/A'
            par_dir[d] = par_dir.get(d, 0) + 1
            if s.date_debut:
                par_mois[s.date_debut.month - 1] += 1

        filename = f"renouvellements_{annee}"

        if format == 'excel':
            wb, ws = make_excel_wb("Renouvellements")
            start = excel_title_block(ws, f"Renouvellements de stage — {annee}",
                                      f"{qs.count()} renouvellement(s)", merge_cols=6)

            excel_header_row(ws, ['Stagiaire', 'Direction', 'Nouveau début', 'Nouvelle fin', 'Ancien stage ID', 'Rémunéré'], row=start)
            for i, s in enumerate(qs, start + 1):
                excel_data_row(ws, [f"{s.prenom} {s.nom}", s.direction,
                    s.date_debut.strftime('%d/%m/%Y') if s.date_debut else '',
                    s.date_fin.strftime('%d/%m/%Y') if s.date_fin else '',
                    s.stage_precedent.id if s.stage_precedent else '',
                    'Oui' if s.remunere else 'Non'], i, even=(i % 2 == 0))
            excel_autofit(ws)

            # Graphique mensuel
            ws_chart = wb.create_sheet("Graphique")
            ws_chart['A1'] = 'Mois'
            ws_chart['B1'] = 'Renouvellements'
            for i, (m, v) in enumerate(zip(NOMS_MOIS_COURT, par_mois), 2):
                ws_chart.cell(i, 1, m)
                ws_chart.cell(i, 2, v)
            excel_add_bar_chart(ws_chart, f"Renouvellements par mois — {annee}",
                                {'min_col': 1, 'min_row': 2, 'max_row': 13},
                                {'min_col': 2, 'min_row': 1, 'max_row': 13},
                                anchor='D1')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer, filename

        # PDF
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, "Renouvellements de stage")
        styles = getSampleStyleSheet()
        elements = pdf_header("Renouvellements de stage",
                              f"Année {annee} · {qs.count()} renouvellement(s)", styles)

        # KPIs
        elements.extend(pdf_section_title("Synthèse", styles))
        elements.append(pdf_kpi_table([
            ('Total renouvellements', qs.count()),
            ('Directions concernées', len(par_dir)),
            ('Mois le plus actif', NOMS_MOIS_COURT[par_mois.index(max(par_mois))] if max(par_mois) > 0 else 'N/A'),
        ]))
        elements.append(Spacer(1, 14))

        # Graphique mensuel
        elements.extend(pdf_section_title("Évolution mensuelle", styles))
        elements.append(make_bar_chart(
            NOMS_MOIS_COURT, par_mois,
            f"Renouvellements par mois — {annee}"
        ))
        elements.append(Spacer(1, 14))

        # Graphique par direction
        if par_dir:
            elements.extend(pdf_section_title("Par direction", styles))
            sorted_dirs = sorted(par_dir.items(), key=lambda x: -x[1])
            elements.append(make_bar_chart(
                [d[0][:20] for d in sorted_dirs], [d[1] for d in sorted_dirs],
                "Renouvellements par direction", horizontal=True
            ))
            elements.append(Spacer(1, 14))

        # Tableau
        elements.extend(pdf_section_title("Liste détaillée", styles))
        data = [['Stagiaire', 'Direction', 'Nouveau début', 'Nouvelle fin', 'Rémunéré']]
        for s in qs:
            data.append([f"{s.prenom} {s.nom}", s.direction,
                s.date_debut.strftime('%d/%m/%Y') if s.date_debut else '',
                s.date_fin.strftime('%d/%m/%Y') if s.date_fin else '',
                'Oui' if s.remunere else 'Non'])
        elements.append(pdf_table(data, [4.5*cm, 3*cm, 2.5*cm, 2.5*cm, 2.5*cm]))
        elements.append(pdf_footer_note(f"Rapport généré automatiquement — CEB — {today.strftime('%d/%m/%Y')}"))
        doc.build(elements)
        buffer.seek(0)
        return buffer, filename

    # ─────────────────────────────────────────────────────────────────────────
    # CONFORMITÉ
    # ─────────────────────────────────────────────────────────────────────────

    @staticmethod
    def generer_audit_actions(params, format='pdf'):
        annee       = int(params.get('annee', date.today().year))
        mois        = params.get('mois', 'tous')
        type_action = params.get('type_action', 'tous')
        today = date.today()

        qs = UserAction.objects.filter(timestamp__year=annee).select_related('user').order_by('-timestamp')
        if mois != 'tous':
            qs = qs.filter(timestamp__month=int(mois))
        if type_action != 'tous':
            qs = qs.filter(action__icontains=type_action)

        total = qs.count()

        # Stats par utilisateur (top 10)
        par_user = {}
        par_mois_data = [0] * 12
        for a in qs[:1000]:
            u = a.user.email if a.user else 'Système'
            par_user[u] = par_user.get(u, 0) + 1
            if a.timestamp:
                par_mois_data[a.timestamp.month - 1] += 1
        top_users = sorted(par_user.items(), key=lambda x: -x[1])[:10]

        filename = f"audit_actions_{annee}"

        if format == 'excel':
            wb, ws = make_excel_wb("Audit")
            start = excel_title_block(ws, f"Audit des actions — {annee}",
                                      f"Total : {total} action(s) · Mois : {'Tous' if mois == 'tous' else NOMS_MOIS[int(mois)]}",
                                      merge_cols=4)

            excel_header_row(ws, ['Date & heure', 'Utilisateur', 'Action'], row=start, fill_color="64748B")
            for i, a in enumerate(qs[:500], start + 1):
                excel_data_row(ws, [
                    a.timestamp.strftime('%d/%m/%Y %H:%M') if a.timestamp else '',
                    a.user.email if a.user else 'Système',
                    a.action,
                ], i, even=(i % 2 == 0))
            excel_autofit(ws)

            # Graphique activité
            if mois == 'tous':
                ws_chart = wb.create_sheet("Graphique Activité")
                ws_chart['A1'] = 'Mois'
                ws_chart['B1'] = 'Actions'
                for i, (m, v) in enumerate(zip(NOMS_MOIS_COURT, par_mois_data), 2):
                    ws_chart.cell(i, 1, m)
                    ws_chart.cell(i, 2, v)
                excel_add_bar_chart(ws_chart, f"Activité mensuelle — {annee}",
                                    {'min_col': 1, 'min_row': 2, 'max_row': 13},
                                    {'min_col': 2, 'min_row': 1, 'max_row': 13},
                                    anchor='D1')

            # Graphique top utilisateurs
            if top_users:
                ws_u = wb.create_sheet("Top Utilisateurs")
                ws_u['A1'] = 'Utilisateur'
                ws_u['B1'] = 'Actions'
                for i, (u, c) in enumerate(top_users, 2):
                    ws_u.cell(i, 1, u[:30])
                    ws_u.cell(i, 2, c)
                excel_add_bar_chart(ws_u, "Top 10 utilisateurs",
                                    {'min_col': 1, 'min_row': 2, 'max_row': len(top_users) + 1},
                                    {'min_col': 2, 'min_row': 1, 'max_row': len(top_users) + 1},
                                    anchor='D1')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer, filename

        # PDF
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, "Audit des actions")
        styles = getSampleStyleSheet()
        elements = pdf_header("Audit des actions utilisateurs",
                              f"Année {annee} · {total} action(s)", styles)

        # KPIs
        elements.extend(pdf_section_title("Synthèse", styles, color='#64748b'))
        elements.append(pdf_kpi_table([
            ('Total actions enregistrées', total),
            ('Utilisateurs distincts', len(par_user)),
            ('Utilisateur le plus actif', top_users[0][0] if top_users else 'N/A', CLR_ACCENT),
        ]))
        elements.append(Spacer(1, 14))

        # Graphique mensuel
        if mois == 'tous':
            elements.extend(pdf_section_title("Activité mensuelle", styles))
            elements.append(make_line_chart(
                NOMS_MOIS_COURT, [('Actions', par_mois_data)],
                f"Évolution de l'activité — {annee}"
            ))
            elements.append(Spacer(1, 14))

        # Graphique top users
        if top_users:
            elements.extend(pdf_section_title("Top 10 utilisateurs", styles))
            elements.append(make_bar_chart(
                [u[0][:20] for u in top_users], [u[1] for u in top_users],
                "Actions par utilisateur", horizontal=True
            ))
            elements.append(Spacer(1, 14))

        # Tableau
        elements.extend(pdf_section_title("Journal des actions", styles))
        data = [['Date & heure', 'Utilisateur', 'Action']]
        for a in qs[:200]:
            data.append([
                a.timestamp.strftime('%d/%m/%Y %H:%M') if a.timestamp else '',
                a.user.email if a.user else 'Système',
                a.action[:80] + '…' if len(a.action) > 80 else a.action,
            ])
        elements.append(pdf_table(data, [3.5*cm, 4.5*cm, 9*cm], header_color=colors.HexColor('#64748b')))
        elements.append(pdf_footer_note(f"Rapport généré automatiquement — CEB — {today.strftime('%d/%m/%Y')}"))
        doc.build(elements)
        buffer.seek(0)
        return buffer, filename

    @staticmethod
    def generer_demandes_archivees(params, format='pdf'):
        annee = int(params.get('annee', date.today().year))
        today = date.today()

        qs = Demande.objects.filter(
            est_archivee=True, date_soumission__year=annee
        ).select_related('etablissement').order_by('-date_soumission')

        # Stats
        total = qs.count()
        par_statut = {}
        par_mois_data = [0] * 12
        for d in qs:
            s = d.statut_stage or 'N/A'
            par_statut[s] = par_statut.get(s, 0) + 1
            if d.date_soumission:
                par_mois_data[d.date_soumission.month - 1] += 1

        filename = f"demandes_archivees_{annee}"

        if format == 'excel':
            wb, ws = make_excel_wb("Archivées")
            start = excel_title_block(ws, f"Demandes archivées — {annee}",
                                      f"{total} dossier(s) archivé(s)", merge_cols=5)

            excel_header_row(ws, ['N° Suivi', 'Demandeur', 'Statut final', 'Date soumission', 'Établissement'], row=start, fill_color="94A3B8")
            for i, d in enumerate(qs, start + 1):
                excel_data_row(ws, [d.tracking_id, f"{d.etudiant_prenom} {d.etudiant_nom}", d.statut_stage,
                    d.date_soumission.strftime('%d/%m/%Y') if d.date_soumission else '',
                    d.etablissement.nom if d.etablissement else ''], i, even=(i % 2 == 0))
            excel_autofit(ws)

            # Graphique statuts
            if par_statut:
                ws_chart = wb.create_sheet("Graphique Statuts")
                ws_chart['A1'] = 'Statut'
                ws_chart['B1'] = 'Nombre'
                for i, (s, c) in enumerate(par_statut.items(), 2):
                    ws_chart.cell(i, 1, s)
                    ws_chart.cell(i, 2, c)
                excel_add_pie_chart(ws_chart, f"Répartition par statut — {annee}",
                                    {'min_col': 1, 'min_row': 1, 'max_row': len(par_statut) + 1},
                                    {'min_col': 2, 'min_row': 1, 'max_row': len(par_statut) + 1},
                                    anchor='D1')

            # Graphique mensuel
            ws_mois = wb.create_sheet("Graphique Mensuel")
            ws_mois['A1'] = 'Mois'
            ws_mois['B1'] = 'Archivées'
            for i, (m, v) in enumerate(zip(NOMS_MOIS_COURT, par_mois_data), 2):
                ws_mois.cell(i, 1, m)
                ws_mois.cell(i, 2, v)
            excel_add_bar_chart(ws_mois, f"Archivages par mois — {annee}",
                                {'min_col': 1, 'min_row': 2, 'max_row': 13},
                                {'min_col': 2, 'min_row': 1, 'max_row': 13},
                                anchor='D1')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer, filename

        # PDF
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, "Demandes archivées")
        styles = getSampleStyleSheet()
        elements = pdf_header("Demandes archivées", f"Année {annee} · {total} dossier(s)", styles)

        # KPIs
        elements.extend(pdf_section_title("Synthèse", styles, color='#94a3b8'))
        elements.append(pdf_kpi_table([
            ('Total archivées', total),
            ('Statuts distincts', len(par_statut)),
        ]))
        elements.append(Spacer(1, 14))

        # Graphique statuts
        if par_statut:
            elements.extend(pdf_section_title("Répartition par statut final", styles))
            elements.append(make_pie_chart(
                list(par_statut.keys()), list(par_statut.values()),
                f"Statuts des dossiers archivés — {annee}"
            ))
            elements.append(Spacer(1, 14))

        # Graphique mensuel
        elements.extend(pdf_section_title("Archivages par mois", styles))
        elements.append(make_bar_chart(
            NOMS_MOIS_COURT, par_mois_data,
            f"Évolution mensuelle des archivages — {annee}"
        ))
        elements.append(Spacer(1, 14))

        # Tableau
        elements.extend(pdf_section_title("Liste des dossiers", styles))
        data = [['N° Suivi', 'Demandeur', 'Statut final', 'Date soumission']]
        for d in qs:
            data.append([d.tracking_id, f"{d.etudiant_prenom} {d.etudiant_nom}", d.statut_stage,
                d.date_soumission.strftime('%d/%m/%Y') if d.date_soumission else ''])
        elements.append(pdf_table(data, [3*cm, 5*cm, 4*cm, 4*cm], header_color=colors.HexColor('#94a3b8')))
        elements.append(pdf_footer_note(f"Rapport généré automatiquement — CEB — {today.strftime('%d/%m/%Y')}"))
        doc.build(elements)
        buffer.seek(0)
        return buffer, filename