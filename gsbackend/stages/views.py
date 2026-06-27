
import logging
from datetime import datetime, timedelta, time
from io import BytesIO
from collections import defaultdict
import pytesseract
import google.generativeai as genai
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer,
    Table, TableStyle, HRFlowable, Image as RLImage,
)
from services.rapport_service import (
    pdf_header, pdf_section_title, pdf_kpi_table, pdf_table, pdf_footer_note,
    make_pdf_doc, make_bar_chart, make_pie_chart, make_line_chart,
    excel_title_block, excel_header_row, excel_data_row, excel_autofit,
    excel_add_bar_chart, excel_add_pie_chart, excel_add_line_chart,
    EXCEL_BORDER, ACCENT_COLOR, HEADER_COLOR, LIGHT_BG, BORDER_COLOR,
    CLR_ACCENT, CLR_SUCCESS, CLR_DANGER, CLR_WARNING, CLR_INFO,
    NOMS_MOIS_COURT,
)
from services.email_service import EmailSenderService

import pandas as pd
from django.conf import settings

import time

from django.db.models import Count
from django.db.models.functions import ExtractMonth, ExtractYear, ExtractWeekDay
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from django.utils.decorators import method_decorator
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from utilisateurs.permissions import HasPermission, PERM_STATS_VIEW, PERM_DATA_EXPORT
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import (
    Stagiaire,Entretien, Demande,UserAction, DemandeAttestation
)

from services.notification_service import NotificationService


import platform
if platform.system() == 'Windows':
    pytesseract.pytesseract.tesseract_cmd = os.getenv('TESSERACT_PATH', r'C:\Program Files\Tesseract-OCR\tesseract.exe')
    POPPLER_PATH = os.getenv('POPPLER_PATH', r'C:\poppler\Library\bin')
else:
    # Chemins Linux/Production
    pytesseract.pytesseract.tesseract_cmd = '/usr/bin/tesseract'
    POPPLER_PATH = '/usr/bin'

import os
from dotenv import load_dotenv

load_dotenv()

logger = logging.getLogger(__name__)

import os
from django.conf import settings
from django.views.decorators.cache import never_cache
import time
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache
from django_ratelimit.decorators import ratelimit
import logging

logger = logging.getLogger(__name__)

@method_decorator([never_cache, ratelimit(key='ip', rate='30/m', method='GET')], name='dispatch')
class DashboardAPIView(APIView):
    permission_classes = [IsAuthenticated]  # ✅ Changé de AllowAny à IsAuthenticated

    def format_date(self, dt: datetime | None) -> str | None:
        """Retourne juste la date au format YYYY-MM-DD"""
        if not dt:
            return None
        return dt.date().isoformat()

    def get(self, request):
        today = datetime.today().date()
        tomorrow = today + timedelta(days=1)

        demandes = Demande.objects.all()
        stages = Stagiaire.objects.all()

        # Entretiens planifiés à venir
        entretiens_a_venir = Entretien.objects.filter(
            date__gte=today,
            status='planifié'
        ).order_by('date', 'heure_debut')[:5]

        return Response({
            "today": today.isoformat(),
            "tomorrow": tomorrow.isoformat(),
            "entretiens_a_venir": [
                {
                    "id": e.id,
                    "titre": e.titre,
                    "date": self.format_date(
                        datetime.combine(e.date, e.heure_debut or time.min)
                    ),
                    "demandeur_id": e.demandeur.id,
                }
                for e in entretiens_a_venir
            ],
            "dernieres_demandes": [
                {
                    "id": d.id,
                    "etudiant_nom": d.etudiant_nom,
                    "etudiant_prenom": d.etudiant_prenom,
                    "statut_stage": d.statut_stage,
                    "etudiant_specialite": d.etudiant_specialite,
                    "date_soumission": d.date_soumission.date().isoformat()
                }
                for d in Demande.objects.filter(statut_stage="En attente").order_by('-date_soumission')[:5]
            ],
            "total_demandes": demandes.count(),
            "stages_en_cours": stages.filter(statut="Actuel").count(),
            "stages_acceptes": demandes.filter(statut_stage="Acceptée").count(),
            "stages_refuses": demandes.filter(statut_stage="Refusée").count(),
            "attestations_total": DemandeAttestation.objects.count(),
            "attestations_en_attente": DemandeAttestation.objects.filter(statut='en_attente').count(),
            "attestations_approuvees": DemandeAttestation.objects.filter(statut='approuvee').count(),
            "attestations_traitees": DemandeAttestation.objects.filter(statut='traitee').count(),
        })
@method_decorator([never_cache, ratelimit(key='user', rate='30/m', method='POST')], name='dispatch')
class MarquerDocumentVuAPI(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request, pk):
        """Marque un document comme vu par l'utilisateur"""
        try:
            demande = get_object_or_404(Demande, pk=pk)
            document_history_id = request.data.get('document_history_id')
            
            if not document_history_id:
                return Response(
                    {'error': 'document_history_id est requis'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Vérifier l'accès à la demande
            if not self.has_access_to_demande(request.user, demande):
                return Response(
                    {'error': 'Accès non autorisé à cette demande'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            succes = demande.mark_document_as_viewed(document_history_id, request.user)
            
            if succes:
                # Enregistrer l'action
                UserAction.objects.create(
                    user=request.user,
                    action=f"Document marqué comme vu pour demande {demande.tracking_id}",
                    performed_by=request.user
                )
                
                return Response({
                    'success': True,
                    'message': 'Document marqué comme vu'
                })
            else:
                return Response({
                    'success': False,
                    'message': 'Document non trouvé'
                }, status=status.HTTP_404_NOT_FOUND)
                
        except Exception as e:
            logger.error(f"Erreur marquage document vu demande {pk}: {str(e)}")
            return Response({
                'success': False,
                'message': 'Erreur lors du marquage du document'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def has_access_to_demande(self, user, demande):
        """Vérifie si l'utilisateur a accès à la demande"""
        return user.is_authenticated


@method_decorator([never_cache, ratelimit(key='user', rate='30/m', method='GET')], name='dispatch')
class PreferencesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        profil = request.user.profil
        data = {
            "loginAlerts": profil.login_alerts,
            "emailSecurity": profil.email_securite,
            "emailUpdates": profil.email_maj,
            "pushLogin": profil.push_connexion,
            "pushSecurity": profil.push_securite,
            "pushUpdates": profil.push_maj,
        }
        return Response(data)

@method_decorator([never_cache, ratelimit(key='user', rate='30/m', method='GET')], name='dispatch')
class UpdatePreferenceAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pref_type):
        profil = request.user.profil
        enabled = request.data.get("enabled", False)

        try:
            # Récupérer l'ancien statut avant modification
            ancien_statut = self._get_ancien_statut(profil, pref_type)
            
            # Mettre à jour la préférence
            message = self._update_preference(profil, pref_type, enabled, request.user)
            
            if not message:
                return Response(
                    {"status": "error", "message": "Type de préférence inconnu"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Sauvegarder les modifications
            profil.save()

            # 🔹 Enregistrer l'action
            self._log_action(request.user, message)

            # 🔔 Notification push
            self._send_push_notification(request.user, message, pref_type, enabled)
           
            return Response({
                "status": "ok", 
                "message": message,
                "preference": pref_type,
                "enabled": enabled
            })

        except Exception as e:
            logger.error(f"❌ Erreur modification préférence {pref_type}: {e}")
            return Response(
                {"status": "error", "message": "Erreur lors de la modification des préférences"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    def _get_ancien_statut(self, profil, pref_type):
        """Récupérer l'ancien statut d'une préférence"""
        mapping = {
            "emailSecurity": profil.email_securite,
            "emailUpdates": profil.email_maj,
            "pushLogin": profil.push_connexion,
            "pushSecurity": profil.push_securite,
            "pushUpdates": profil.push_maj,
            "loginAlerts": profil.login_alerts,
        }
        return mapping.get(pref_type, False)

    def _update_preference(self, profil, pref_type, enabled, user):
        """Met à jour la préférence et retourne le message"""
        
        if pref_type == "emailSecurity":
            profil.email_securite = enabled
            return f"Emails de sécurité {'activés' if enabled else 'désactivés'}"
            
        elif pref_type == "emailUpdates":
            # ✅ Email de bienvenue si activation des nouveautés
            if enabled and not profil.email_maj:
                self._send_newsletter_welcome_email(user)
            profil.email_maj = enabled
            return f"Emails des nouveautés {'activés' if enabled else 'désactivés'}"
            
        elif pref_type == "pushLogin":
            profil.push_connexion = enabled
            return f"Notifications push de connexion {'activées' if enabled else 'désactivées'}"
            
        elif pref_type == "pushSecurity":
            profil.push_securite = enabled
            return f"Notifications push de sécurité {'activées' if enabled else 'désactivées'}"
            
        elif pref_type == "pushUpdates":
            profil.push_maj = enabled
            return f"Notifications push des mises à jour {'activées' if enabled else 'désactivées'}"
            
        elif pref_type == "loginAlerts":
            profil.login_alerts = enabled
            return f"Alertes de connexion {'activées' if enabled else 'désactivées'}"
        
        return None

    def _log_action(self, user, message):
        """Enregistre l'action dans l'historique"""
        try:
            UserAction.objects.create(
                user=user,
                action=f"Préférence modifiée: {message}",
                performed_by=user
            )
            logger.info(f"✅ Action enregistrée pour {user.email}")
        except Exception as e:
            logger.error(f"❌ Erreur enregistrement action: {e}")

    def _send_push_notification(self, user, message, pref_type, enabled):
        """Envoie une notification push"""
        try:
            NotificationService.notifier_changement_preference(
                user=user,
                preference=pref_type,
                message=message,
                enabled=enabled
            )
            logger.info(f"✅ Notification push envoyée pour {pref_type}")
        except Exception as e:
            logger.error(f"❌ Erreur notification push: {e}")

    def _send_newsletter_welcome_email(self, user):
        """
        ✅ Email de bienvenue pour abonnement aux nouveautés
        Utilise le service centralisé
        """
        try:
            success = EmailSenderService.send_newsletter_welcome(
                user=user,
                async_send=True
            )
            
            if success:
                logger.info(f"✅ Email bienvenue nouveautés envoyé à {user.email}")
        except Exception as e:
            logger.error(f"❌ Erreur email bienvenue nouveautés: {e}")

    def _send_confirmation_email(self, user, profil, message, pref_type, enabled, ancien_statut):
        """
        ✅ Email de confirmation de changement de préférence
        Utilise le service centralisé
        """
        # Ne pas envoyer d'email si on désactive les emails de sécurité
        if pref_type == "emailSecurity":
            return
        
        # Envoyer uniquement si les emails de sécurité sont activés
        if not profil.email_securite:
            return
        
        try:
            success = EmailSenderService.send_preference_changed(
                user=user,
                preference_name=self._get_nom_preference(pref_type),
                preference_type=pref_type,
                enabled=enabled,
                description=self._get_description_impact(pref_type, enabled),
                async_send=True
            )
            
            if success:
                logger.info(f"✅ Email confirmation préférence envoyé à {user.email}")
        except Exception as e:
            logger.error(f"❌ Erreur email confirmation: {e}")

    def _get_nom_preference(self, pref_type):
        """Nom lisible de la préférence"""
        noms = {
            "emailSecurity": "Emails de sécurité",
            "emailUpdates": "Emails des nouveautés",
            "pushLogin": "Notifications push de connexion",
            "pushSecurity": "Notifications push de sécurité",
            "pushUpdates": "Notifications push des mises à jour",
            "loginAlerts": "Alertes de connexion",
        }
        return noms.get(pref_type, "Préférence inconnue")

    def _get_description_impact(self, pref_type, enabled):
        """Description de l'impact de la modification"""
        descriptions = {
            "emailSecurity": {
                True: "Vous recevrez désormais des emails pour les actions importantes de sécurité sur votre compte.",
                False: "Vous ne recevrez plus d'emails pour les actions de sécurité. Les notifications push restent actives."
            },
            "emailUpdates": {
                True: "Vous recevrez désormais des emails concernant les nouvelles fonctionnalités et mises à jour de la plateforme.",
                False: "Vous ne recevrez plus d'emails concernant les nouveautés de la plateforme."
            },
            "pushLogin": {
                True: "Vous recevrez désormais des notifications push pour chaque nouvelle connexion à votre compte.",
                False: "Vous ne recevrez plus de notifications push pour les nouvelles connexions."
            },
            "pushSecurity": {
                True: "Vous recevrez désormais des notifications push pour les alertes de sécurité importantes.",
                False: "Vous ne recevrez plus de notifications push pour les alertes de sécurité."
            },
            "pushUpdates": {
                True: "Vous recevrez désormais des notifications push pour les mises à jour de la plateforme.",
                False: "Vous ne recevrez plus de notifications push pour les mises à jour."
            },
            "loginAlerts": {
                True: "Vous serez alerté pour chaque nouvelle connexion détectée sur votre compte.",
                False: "Vous ne serez plus alerté pour les nouvelles connexions."
            }
        }
        return descriptions.get(pref_type, {}).get(enabled, "")


def to_date(value):
    if isinstance(value, datetime):
        return value.date()
    return value

@method_decorator([never_cache, ratelimit(key='user', rate='30/m', method='GET')], name='dispatch')
class StatistiquesAPIView(APIView):
    permission_classes = [IsAuthenticated, HasPermission(PERM_STATS_VIEW)]

    def get(self, request):
        current_year = datetime.today().year
        
        # Gestion des intervalles d'années
        annee_debut = request.GET.get("annee_debut")
        annee_fin = request.GET.get("annee_fin")
        
        if annee_debut and annee_fin:
            try:
                annee_debut = int(annee_debut)
                annee_fin = int(annee_fin)
                intervalle_annees = True
            except ValueError:
                annee_debut = current_year - 4
                annee_fin = current_year
                intervalle_annees = True
        else:
            try:
                annee_cible = int(request.GET.get("annee", current_year))
                intervalle_annees = False
            except ValueError:
                annee_cible = current_year
                intervalle_annees = False

        if intervalle_annees:
            return self.get_stats_intervalle(annee_debut, annee_fin)
        else:
            return self.get_stats_annuelle(annee_cible)

    def get_stats_annuelle(self, annee_cible):
        """Récupère les statistiques pour une année spécifique"""
        # 1. Total des demandes
        total_demandes = Demande.objects.filter(date_soumission__year=annee_cible).count()

        # 2. Taux d'acceptation (basé sur les décisions finales)
        demandes_accept = Demande.objects.filter(
            statut_stage="Acceptée", date_soumission__year=annee_cible
        ).count()
        demandes_refuse = Demande.objects.filter(
            statut_stage="Refusée", date_soumission__year=annee_cible
        ).count()
        demandes_traitees = demandes_accept + demandes_refuse
        taux_acceptation = round((demandes_accept / demandes_traitees) * 100, 2) if demandes_traitees else 0

        # 3. Délai moyen de traitement (pour les demandes traitées)
        delais = []
        stagiaires_acceptes = Stagiaire.objects.filter(
            demande__statut_stage__in=["Acceptée", "Refusée"],
            demande__date_soumission__year=annee_cible,
            demande__date_soumission__isnull=False,
            date_accord__isnull=False,
        ).values("demande__date_soumission", "date_accord", "demande__statut_stage")

        for s in stagiaires_acceptes:
            d_soumission = to_date(s["demande__date_soumission"])
            d_accord = to_date(s["date_accord"])
            if d_soumission and d_accord:
                delta_days = (d_accord - d_soumission).days
                delais.append(max(delta_days, 0))

        delai_moyen = round(sum(delais) / len(delais), 1) if delais else 0

        # 4. Durée moyenne de stage
        durees_stage = []
        stages = Stagiaire.objects.filter(
            demande__date_soumission__year=annee_cible,
            date_debut__isnull=False,
            date_fin__isnull=False,
        ).values("date_debut", "date_fin")

        for st in stages:
            date_debut = to_date(st["date_debut"])
            date_fin = to_date(st["date_fin"])
            if date_debut and date_fin and date_fin >= date_debut:
                durees_stage.append((date_fin - date_debut).days)

        duree_moyenne_stage = round(sum(durees_stage) / len(durees_stage), 1) if durees_stage else 0

        # 5. Stages en cours (statut "Actuel")
        stages_en_cours = Stagiaire.objects.filter(
            statut="Actuel", demande__date_soumission__year=annee_cible
        ).count()

        # 6. Demandes par mois
        demandes_par_mois_qs = (
            Demande.objects.filter(date_soumission__year=annee_cible)
            .annotate(mois=ExtractMonth("date_soumission"))
            .values("mois")
            .annotate(count=Count("id"))
            .order_by("mois")
        )
        demandes_par_mois = [0] * 12
        for item in demandes_par_mois_qs:
            demandes_par_mois[item["mois"] - 1] = item["count"]

        # 7. Répartition des statuts (incluant "En cours de traitement")
        repartition_qs = (
            Demande.objects.filter(date_soumission__year=annee_cible)
            .values("statut_stage")
            .annotate(count=Count("id"))
        )
        
        # Calcul des pourcentages
        total_annee = sum(item["count"] for item in repartition_qs) or 1
        repartition_statuts = {
            "Acceptée": 0, 
            "Refusée": 0, 
            "En attente": 0,
            "En cours de traitement": 0
        }
        
        for item in repartition_qs:
            statut = item["statut_stage"]
            if statut in repartition_statuts:
                repartition_statuts[statut] = round((item["count"] / total_annee) * 100, 1)

        # 8. Performance hebdomadaire
        perf_hebdo_qs = (
            Demande.objects.filter(date_soumission__year=annee_cible)
            .annotate(jour_sem=ExtractWeekDay("date_soumission"))
            .values("jour_sem")
            .annotate(count=Count("id"))
        )
        perf_hebdo = [0] * 7
        for item in perf_hebdo_qs:
            jour = item["jour_sem"]  # 1=dimanche
            idx = (jour + 5) % 7  # lundi=0
            perf_hebdo[idx] = item["count"]

        # 9. Demandes en cours de traitement (pour indicateur spécifique)
        demandes_en_cours = Demande.objects.filter(
            statut_stage="En cours de traitement",
            date_soumission__year=annee_cible
        ).count()

        # 10. Comparaison annuelle (demandes acceptées)
        annee_min = 2020
        comp_annee_qs = (
            Demande.objects.filter(statut_stage="Acceptée", date_soumission__year__gte=annee_min)
            .annotate(annee=ExtractYear("date_soumission"))
            .values("annee")
            .annotate(count=Count("id"))
            .order_by("annee")
        )
        comp_annee = defaultdict(int)
        for item in comp_annee_qs:
            comp_annee[item["annee"]] = item["count"]

        annees_range = list(range(annee_min, datetime.today().year + 1))
        comp_annee_list = [comp_annee[annee] for annee in annees_range]

        return Response(
            {
                "type": "annuelle",
                "annee_cible": annee_cible,
                "total_demandes": total_demandes,
                "taux_acceptation": taux_acceptation,
                "delai_moyen": delai_moyen,
                "duree_moyenne_stage": duree_moyenne_stage,
                "stages_en_cours": stages_en_cours,
                "demandes_en_cours_traitement": demandes_en_cours,
                "demandes_par_mois": demandes_par_mois,
                "repartition_statuts": [
                    repartition_statuts["Acceptée"],
                    repartition_statuts["Refusée"],
                    repartition_statuts["En attente"],
                    repartition_statuts["En cours de traitement"],
                ],
                "perf_hebdo": perf_hebdo,
                "comp_annee": comp_annee_list,
                "comp_annee_labels": annees_range,
            }
        )

    def get_stats_intervalle(self, annee_debut, annee_fin):
        """Récupère les statistiques pour un intervalle d'années"""
        annees = list(range(annee_debut, annee_fin + 1))
        
        # Statistiques par année
        stats_par_annee = []
        for annee in annees:
            total_demandes = Demande.objects.filter(date_soumission__year=annee).count()
            demandes_acceptees = Demande.objects.filter(
                statut_stage="Acceptée", date_soumission__year=annee
            ).count()
            demandes_refusees = Demande.objects.filter(
                statut_stage="Refusée", date_soumission__year=annee
            ).count()
            demandes_en_cours = Demande.objects.filter(
                statut_stage="En cours de traitement", date_soumission__year=annee
            ).count()
            demandes_traitees = demandes_acceptees + demandes_refusees
            taux_acceptation = round((demandes_acceptees / demandes_traitees * 100), 2) if demandes_traitees else 0
            
            stats_par_annee.append({
                "annee": annee,
                "total_demandes": total_demandes,
                "demandes_acceptees": demandes_acceptees,
                "demandes_refusees": demandes_refusees,
                "demandes_en_cours": demandes_en_cours,
                "taux_acceptation": taux_acceptation
            })

        # Évolution mensuelle sur l'intervalle
        evolution_mensuelle = []
        for annee in annees:
            demandes_mois = [0] * 12
            demandes_mois_qs = (
                Demande.objects.filter(date_soumission__year=annee)
                .annotate(mois=ExtractMonth("date_soumission"))
                .values("mois")
                .annotate(count=Count("id"))
                .order_by("mois")
            )
            for item in demandes_mois_qs:
                demandes_mois[item["mois"] - 1] = item["count"]
            evolution_mensuelle.append({
                "annee": annee,
                "demandes_par_mois": demandes_mois
            })

        # Statistiques consolidées
        total_intervalle = sum(item["total_demandes"] for item in stats_par_annee)
        total_acceptees_intervalle = sum(item["demandes_acceptees"] for item in stats_par_annee)
        total_refusees_intervalle = sum(item["demandes_refusees"] for item in stats_par_annee)
        total_en_cours_intervalle = sum(item["demandes_en_cours"] for item in stats_par_annee)
        total_traitees_intervalle = total_acceptees_intervalle + total_refusees_intervalle
        taux_acceptation_intervalle = round((total_acceptees_intervalle / total_traitees_intervalle * 100), 2) if total_traitees_intervalle else 0

        # AJOUT : Données pour compatibilité avec le frontend
        # Calcul du délai moyen sur tout l'intervalle
        delais = []
        stagiaires_acceptes = Stagiaire.objects.filter(
            demande__statut_stage__in=["Acceptée", "Refusée"],
            demande__date_soumission__year__gte=annee_debut,
            demande__date_soumission__year__lte=annee_fin,
            demande__date_soumission__isnull=False,
            date_accord__isnull=False,
        ).values("demande__date_soumission", "date_accord")

        for s in stagiaires_acceptes:
            d_soumission = to_date(s["demande__date_soumission"])
            d_accord = to_date(s["date_accord"])
            if d_soumission and d_accord:
                delta_days = (d_accord - d_soumission).days
                delais.append(max(delta_days, 0))

        delai_moyen = round(sum(delais) / len(delais), 1) if delais else 0

        # Durée moyenne de stage sur l'intervalle
        durees_stage = []
        stages = Stagiaire.objects.filter(
            demande__date_soumission__year__gte=annee_debut,
            demande__date_soumission__year__lte=annee_fin,
            date_debut__isnull=False,
            date_fin__isnull=False,
        ).values("date_debut", "date_fin")

        for st in stages:
            date_debut = to_date(st["date_debut"])
            date_fin = to_date(st["date_fin"])
            if date_debut and date_fin and date_fin >= date_debut:
                durees_stage.append((date_fin - date_debut).days)

        duree_moyenne_stage = round(sum(durees_stage) / len(durees_stage), 1) if durees_stage else 0

        # Stages actuellement en cours dans l'intervalle
        stages_en_cours = Stagiaire.objects.filter(
            statut="Actuel",
            demande__date_soumission__year__gte=annee_debut,
            demande__date_soumission__year__lte=annee_fin
        ).count()

        return Response({
            "type": "intervalle",
            "annee_debut": annee_debut,
            "annee_fin": annee_fin,
            "stats_par_annee": stats_par_annee,
            "evolution_mensuelle": evolution_mensuelle,
            "total_intervalle": total_intervalle,
            "total_acceptees_intervalle": total_acceptees_intervalle,
            "total_refusees_intervalle": total_refusees_intervalle,
            "total_en_cours_intervalle": total_en_cours_intervalle,
            "taux_acceptation_intervalle": taux_acceptation_intervalle,
            "annees_labels": annees,
            # Données supplémentaires pour compatibilité
            "total_demandes": total_intervalle,
            "taux_acceptation": taux_acceptation_intervalle,
            "delai_moyen": delai_moyen,
            "duree_moyenne_stage": duree_moyenne_stage,
            "stages_en_cours": stages_en_cours,
            "demandes_en_cours_traitement": total_en_cours_intervalle,
            # Valeurs par défaut pour éviter les erreurs
            "demandes_par_mois": [0] * 12,
            "repartition_statuts": [0, 0, 0, 0],
            "perf_hebdo": [0] * 7,
            "comp_annee": [],
            "comp_annee_labels": [],
        })

@method_decorator([never_cache, ratelimit(key='user', rate='30/m', method='GET')], name='dispatch')
class ExportStatsAPIView(APIView):
    permission_classes = [IsAuthenticated, HasPermission(PERM_DATA_EXPORT)]

    def get(self, request, format_type):
        annee_cible = int(request.GET.get("annee", datetime.today().year))
        annee_debut = request.GET.get("annee_debut")
        annee_fin = request.GET.get("annee_fin")
        report_type = request.GET.get("type", "general")

        if annee_debut and annee_fin:
            stats_data = self.get_stats_intervalle_data(int(annee_debut), int(annee_fin))
            is_intervalle = True
        else:
            stats_data = self.get_stats_annuelle_data(annee_cible)
            is_intervalle = False

        if format_type == 'excel':
            return self.export_excel(stats_data, annee_cible, report_type, is_intervalle)
        elif format_type == 'pdf':
            return self.export_pdf(stats_data, annee_cible, report_type, is_intervalle)
        else:
            return Response({"error": "Format non supporté"}, status=400)

    def get_stats_annuelle_data(self, annee_cible):
        view = StatistiquesAPIView()
        view.request = self.request
        response = view.get(self.request)
        return response.data

    def get_stats_intervalle_data(self, annee_debut, annee_fin):
        view = StatistiquesAPIView()
        view.request = self.request
        response = view.get(self.request)
        return response.data

    def export_excel(self, stats_data, annee, report_type, is_intervalle=False):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, PieChart, LineChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.series import DataPoint

        wb = openpyxl.Workbook()
        ws = wb.active

        if is_intervalle:
            # ── Feuille Synthèse ──
            ws.title = "Synthèse Intervalle"
            ws.sheet_properties.tabColor = '6366F1'
            start = excel_title_block(ws,
                f"Rapport des Stages {stats_data['annee_debut']} - {stats_data['annee_fin']}",
                f"Synthèse de la période", merge_cols=2)

            excel_header_row(ws, ['Métrique', 'Valeur'], row=start)
            kpis = [
                ("Période analysée", f"{stats_data['annee_debut']} - {stats_data['annee_fin']}"),
                ("Total des demandes", stats_data["total_intervalle"]),
                ("Demandes acceptées", stats_data["total_acceptees_intervalle"]),
                ("Demandes refusées", stats_data["total_refusees_intervalle"]),
                ("Demandes en cours", stats_data["total_en_cours_intervalle"]),
                ("Taux d'acceptation", f"{stats_data['taux_acceptation_intervalle']}%"),
            ]
            for i, (k, v) in enumerate(kpis, start + 1):
                excel_data_row(ws, [k, v], i, even=(i % 2 == 0))
            excel_autofit(ws)

            # ── Feuille Détail par Année ──
            ws2 = wb.create_sheet("Détail par Année")
            ws2.sheet_properties.tabColor = '10B981'
            headers = ["Année", "Total", "Acceptées", "Refusées", "En Cours", "Taux"]
            start2 = excel_title_block(ws2, "Détail par année", merge_cols=6)
            excel_header_row(ws2, headers, row=start2)
            for i, stat in enumerate(stats_data["stats_par_annee"], start2 + 1):
                excel_data_row(ws2, [
                    stat["annee"], stat["total_demandes"], stat["demandes_acceptees"],
                    stat["demandes_refusees"], stat["demandes_en_cours"],
                    f"{stat['taux_acceptation']}%"
                ], i, even=(i % 2 == 0))
            excel_autofit(ws2)

            # Graphique barres par année
            n_annees = len(stats_data["stats_par_annee"])
            if n_annees > 0:
                chart = BarChart()
                chart.type = "col"
                chart.title = "Évolution par année"
                chart.width = 20
                chart.height = 12
                d1 = Reference(ws2, min_col=3, min_row=start2, max_row=start2 + n_annees)
                d2 = Reference(ws2, min_col=4, min_row=start2, max_row=start2 + n_annees)
                d3 = Reference(ws2, min_col=5, min_row=start2, max_row=start2 + n_annees)
                cats = Reference(ws2, min_col=1, min_row=start2 + 1, max_row=start2 + n_annees)
                chart.add_data(d1, titles_from_data=True)
                chart.add_data(d2, titles_from_data=True)
                chart.add_data(d3, titles_from_data=True)
                chart.set_categories(cats)
                chart.series[0].graphicalProperties.solidFill = '10B981'
                chart.series[1].graphicalProperties.solidFill = 'EF4444'
                chart.series[2].graphicalProperties.solidFill = 'F59E0B'
                ws2.add_chart(chart, f"H{start2}")

            # ── Feuille Évolution Mensuelle ──
            if "evolution_mensuelle" in stats_data and stats_data["evolution_mensuelle"]:
                ws3 = wb.create_sheet("Évolution Mensuelle")
                ws3.sheet_properties.tabColor = 'F59E0B'
                months = ["Jan", "Fév", "Mar", "Avr", "Mai", "Jun", "Jul", "Aoû", "Sep", "Oct", "Nov", "Déc"]
                annee_cols = [str(e["annee"]) for e in stats_data["evolution_mensuelle"]]
                all_headers = ["Mois"] + annee_cols
                excel_header_row(ws3, all_headers, row=1)
                for i, m in enumerate(months):
                    row_vals = [m]
                    for e in stats_data["evolution_mensuelle"]:
                        row_vals.append(e["demandes_par_mois"][i])
                    excel_data_row(ws3, row_vals, i + 2, even=(i % 2 == 0))
                excel_autofit(ws3)

                # Line chart évolution
                chart_line = LineChart()
                chart_line.title = "Évolution mensuelle par année"
                chart_line.width = 20
                chart_line.height = 12
                cats = Reference(ws3, min_col=1, min_row=2, max_row=13)
                line_colors = ['6366F1', '10B981', 'F59E0B', 'EF4444', '3B82F6']
                for col_idx in range(2, len(annee_cols) + 2):
                    data = Reference(ws3, min_col=col_idx, min_row=1, max_row=13)
                    chart_line.add_data(data, titles_from_data=True)
                chart_line.set_categories(cats)
                for i, s in enumerate(chart_line.series):
                    s.graphicalProperties.line.solidFill = line_colors[i % len(line_colors)]
                    s.graphicalProperties.line.width = 25000
                    s.smooth = True
                ws3.add_chart(chart_line, "A16")

        else:
            # ── Année unique ──

            # Feuille 1: Statistiques générales
            ws.title = "Statistiques Générales"
            ws.sheet_properties.tabColor = '6366F1'
            start = excel_title_block(ws,
                f"Rapport des Stages — {annee}",
                f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", merge_cols=2)

            excel_header_row(ws, ['Métrique', 'Valeur'], row=start)
            kpis = [
                ("Année analysée", stats_data["annee_cible"]),
                ("Total des demandes", stats_data["total_demandes"]),
                ("Taux d'acceptation", f"{stats_data['taux_acceptation']}%"),
                ("Délai moyen de traitement", f"{stats_data['delai_moyen']} jours"),
                ("Durée moyenne de stage", f"{stats_data['duree_moyenne_stage']} jours"),
                ("Stages en cours", stats_data["stages_en_cours"]),
                ("Demandes en cours de traitement", stats_data["demandes_en_cours_traitement"]),
            ]
            for i, (k, v) in enumerate(kpis, start + 1):
                excel_data_row(ws, [k, v], i, even=(i % 2 == 0))
            excel_autofit(ws)

            # Feuille 2: Demandes par mois + graphique
            ws_mois = wb.create_sheet("Demandes par Mois")
            ws_mois.sheet_properties.tabColor = '10B981'
            months = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                       "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            excel_header_row(ws_mois, ['Mois', 'Nombre de demandes'], row=1)
            for i, (m, v) in enumerate(zip(months, stats_data["demandes_par_mois"]), 2):
                excel_data_row(ws_mois, [m, v], i, even=(i % 2 == 0))
            excel_autofit(ws_mois)

            excel_add_bar_chart(ws_mois, f"Demandes par mois — {annee}",
                                {'min_col': 1, 'min_row': 2, 'max_row': 13},
                                {'min_col': 2, 'min_row': 1, 'max_row': 13},
                                anchor='D1', width=20)

            # Feuille 3: Répartition par statut + pie chart
            ws_statut = wb.create_sheet("Répartition Statuts")
            ws_statut.sheet_properties.tabColor = 'F59E0B'
            statuts = ["Acceptées", "Refusées", "En attente", "En cours de traitement"]
            pourcentages = stats_data["repartition_statuts"]
            total_demandes = stats_data["total_demandes"]

            excel_header_row(ws_statut, ['Statut', 'Pourcentage', 'Nombre'], row=1)
            for i, (s, p) in enumerate(zip(statuts, pourcentages), 2):
                nombre = int(total_demandes * p / 100)
                excel_data_row(ws_statut, [s, f"{p}%", nombre], i, even=(i % 2 == 0))
            excel_autofit(ws_statut)

            excel_add_pie_chart(ws_statut, f"Répartition par statut — {annee}",
                                {'min_col': 1, 'min_row': 1, 'max_row': 5},
                                {'min_col': 3, 'min_row': 1, 'max_row': 5},
                                anchor='E1')

            # Feuille 4: Performance hebdomadaire + graphique
            ws_hebdo = wb.create_sheet("Performance Hebdo")
            ws_hebdo.sheet_properties.tabColor = 'EF4444'
            jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
            excel_header_row(ws_hebdo, ['Jour', 'Nombre de demandes'], row=1)
            for i, (j, v) in enumerate(zip(jours, stats_data["perf_hebdo"]), 2):
                excel_data_row(ws_hebdo, [j, v], i, even=(i % 2 == 0))
            excel_autofit(ws_hebdo)

            excel_add_bar_chart(ws_hebdo, "Performance hebdomadaire",
                                {'min_col': 1, 'min_row': 2, 'max_row': 8},
                                {'min_col': 2, 'min_row': 1, 'max_row': 8},
                                anchor='D1')

            # Feuille 5: Comparaison annuelle + graphique
            ws_comp = wb.create_sheet("Comparaison Annuelle")
            ws_comp.sheet_properties.tabColor = '3B82F6'
            excel_header_row(ws_comp, ['Année', 'Demandes acceptées'], row=1)
            comp_labels = stats_data["comp_annee_labels"]
            comp_vals = stats_data["comp_annee"]
            for i, (a, v) in enumerate(zip(comp_labels, comp_vals), 2):
                excel_data_row(ws_comp, [a, v], i, even=(i % 2 == 0))
            excel_autofit(ws_comp)

            n_comp = len(comp_labels)
            if n_comp > 0:
                excel_add_bar_chart(ws_comp, "Comparaison annuelle",
                                    {'min_col': 1, 'min_row': 2, 'max_row': n_comp + 1},
                                    {'min_col': 2, 'min_row': 1, 'max_row': n_comp + 1},
                                    anchor='D1')

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"rapport_stages_{stats_data['annee_debut']}_{stats_data['annee_fin']}.xlsx" if is_intervalle else f"rapport_stages_{annee}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def export_pdf(self, stats_data, annee, report_type, is_intervalle=False):
        buffer = BytesIO()
        doc = make_pdf_doc(buffer, f"Rapport des Stages — {annee}")
        styles = getSampleStyleSheet()
        story = []

        if is_intervalle:
            story.extend(pdf_header(
                f"Rapport des Stages {stats_data['annee_debut']} - {stats_data['annee_fin']}",
                "Analyse comparative pluriannuelle", styles))

            # KPIs
            story.extend(pdf_section_title("Synthèse de la période", styles))
            story.append(pdf_kpi_table([
                ('Période', f"{stats_data['annee_debut']} - {stats_data['annee_fin']}"),
                ('Total des demandes', stats_data["total_intervalle"]),
                ('Demandes acceptées', stats_data["total_acceptees_intervalle"], CLR_SUCCESS),
                ('Demandes refusées', stats_data["total_refusees_intervalle"], CLR_DANGER),
                ('Demandes en cours', stats_data["total_en_cours_intervalle"], CLR_WARNING),
                ("Taux d'acceptation", f"{stats_data['taux_acceptation_intervalle']}%"),
            ]))
            story.append(Spacer(1, 16))

            # Tableau détail par année
            story.extend(pdf_section_title("Détail par année", styles))
            data_annees = [['Année', 'Total', 'Acceptées', 'Refusées', 'En cours', 'Taux']]
            annee_labels = []
            annee_totals = []
            annee_acc = []
            for stat in stats_data["stats_par_annee"]:
                data_annees.append([
                    str(stat["annee"]), str(stat["total_demandes"]),
                    str(stat["demandes_acceptees"]), str(stat["demandes_refusees"]),
                    str(stat["demandes_en_cours"]), f"{stat['taux_acceptation']}%"
                ])
                annee_labels.append(str(stat["annee"]))
                annee_totals.append(stat["total_demandes"])
                annee_acc.append(stat["demandes_acceptees"])
            story.append(pdf_table(data_annees,
                                   [2*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm]))
            story.append(Spacer(1, 16))

            # Graphique barres par année
            if annee_labels:
                story.extend(pdf_section_title("Évolution par année", styles))
                story.append(make_bar_chart(
                    annee_labels, annee_totals,
                    "Total des demandes par année"
                ))
                story.append(Spacer(1, 16))

            # Graphique évolution mensuelle multi-année
            if "evolution_mensuelle" in stats_data and stats_data["evolution_mensuelle"]:
                story.extend(pdf_section_title("Évolution mensuelle", styles))
                datasets = []
                for e in stats_data["evolution_mensuelle"]:
                    datasets.append((str(e["annee"]), e["demandes_par_mois"]))
                story.append(make_line_chart(
                    NOMS_MOIS_COURT, datasets,
                    "Tendance mensuelle par année"
                ))

        else:
            # ── Année unique ──
            story.extend(pdf_header(
                f"Rapport des Stages — {annee}",
                f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles))

            # KPIs
            story.extend(pdf_section_title("Statistiques générales", styles))
            story.append(pdf_kpi_table([
                ('Année analysée', stats_data["annee_cible"]),
                ('Total des demandes', stats_data["total_demandes"]),
                ("Taux d'acceptation", f"{stats_data['taux_acceptation']}%", CLR_SUCCESS),
                ('Délai moyen de traitement', f"{stats_data['delai_moyen']} jours"),
                ('Durée moyenne de stage', f"{stats_data['duree_moyenne_stage']} jours"),
                ('Stages en cours', stats_data["stages_en_cours"]),
                ('Demandes en cours', stats_data["demandes_en_cours_traitement"]),
            ]))
            story.append(Spacer(1, 16))

            # Répartition par statut — tableau + pie chart
            statuts = ["Acceptées", "Refusées", "En attente", "En cours de traitement"]
            pourcentages = stats_data["repartition_statuts"]
            total_demandes = stats_data["total_demandes"]
            nombres = [int(total_demandes * p / 100) for p in pourcentages]

            story.extend(pdf_section_title("Répartition par statut", styles))
            data_statuts = [['Statut', 'Pourcentage', 'Nombre']]
            for s, p, n in zip(statuts, pourcentages, nombres):
                data_statuts.append([s, f"{p}%", str(n)])
            story.append(pdf_table(data_statuts, [5*cm, 4*cm, 4*cm]))
            story.append(Spacer(1, 10))

            story.append(make_pie_chart(statuts, nombres,
                                        f"Répartition par statut — {annee}"))
            story.append(Spacer(1, 16))

            # Demandes par mois — tableau + line chart
            months_fr = NOMS_MOIS_COURT
            demandes_mois = stats_data["demandes_par_mois"]

            story.extend(pdf_section_title("Demandes par mois", styles))
            story.append(make_line_chart(
                months_fr, [('Demandes', demandes_mois)],
                f"Évolution mensuelle — {annee}"
            ))
            story.append(Spacer(1, 8))

            data_mois = [['Mois', 'Demandes']]
            for m, v in zip(months_fr, demandes_mois):
                data_mois.append([m, str(v)])
            story.append(pdf_table(data_mois, [5*cm, 5*cm],
                                   header_color=colors.HexColor('#10b981')))
            story.append(Spacer(1, 16))

            # Performance hebdomadaire — bar chart
            jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
            perf = stats_data["perf_hebdo"]

            story.extend(pdf_section_title("Performance hebdomadaire", styles))
            story.append(make_bar_chart(jours, perf,
                                        "Demandes par jour de la semaine"))
            story.append(Spacer(1, 16))

            # Comparaison annuelle — bar chart
            comp_labels = stats_data["comp_annee_labels"]
            comp_vals = stats_data["comp_annee"]
            if comp_labels:
                story.extend(pdf_section_title("Comparaison annuelle", styles))
                story.append(make_bar_chart(
                    [str(a) for a in comp_labels], comp_vals,
                    "Demandes acceptées par année"
                ))

        story.append(pdf_footer_note(
            f"Rapport généré automatiquement — CEB Gestion des Stages — {datetime.now().strftime('%d/%m/%Y')}"))

        doc.build(story)
        buffer.seek(0)

        filename = f"rapport_stages_{stats_data['annee_debut']}_{stats_data['annee_fin']}.pdf" if is_intervalle else f"rapport_stages_{annee}.pdf"

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

class SidebarCountsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        demandes_attente = Demande.objects.filter(statut_stage='En attente', est_archivee=False).count()
        demandes_acceptees = Demande.objects.filter(statut_stage='Acceptée').count()
        demandes_refusees = Demande.objects.filter(statut_stage='Refusée').count()
        demandes_traitement = Demande.objects.filter(statut_stage='En cours de traitement').count()
        demandes_en_acceptation = Demande.objects.filter(statut_stage='Pré-acceptée').count()

        demandes_archivees = Demande.objects.filter(est_archivee=True).count()


        stages_a_venir = Stagiaire.objects.filter(statut='À venir').count()
        stages_en_cours = Stagiaire.objects.filter(statut='Actuel').count()
        stages_termines = Stagiaire.objects.filter(statut='Terminé', date_cloture__isnull=True).count()
        stages_clotures = Stagiaire.objects.filter(statut='Terminé', date_cloture__isnull=False).count()

        demandes_attestation = DemandeAttestation.objects.all().count()
        attestations_en_attente = DemandeAttestation.objects.filter(statut='en_attente').count()
        attestations_approuvees = DemandeAttestation.objects.filter(statut='approuvee').count()
        attestations_refusees = DemandeAttestation.objects.filter(statut='refusee').count()
        attestations_traitees = DemandeAttestation.objects.filter(statut='traitee').count()

        return Response({
            'count_demandes_attente': demandes_attente,
            'count_demandes_acceptees': demandes_acceptees,
            'count_demandes_refusees': demandes_refusees,
            'count_demandes_traitement': demandes_traitement,
            'count_demandes_en_acceptation': demandes_en_acceptation,
            'count_demandes_archivees': demandes_archivees,
            'count_stages_avenir': stages_a_venir,
            'count_stages_en_cours': stages_en_cours,
            'count_stages_termines': stages_termines,
            'count_stages_clotures': stages_clotures,
            'count_demandes_attestation': demandes_attestation,
            'count_attestations_en_attente': attestations_en_attente,
            'count_attestations_approuvees': attestations_approuvees,
            'count_attestations_refusees': attestations_refusees,
            'count_attestations_traitees': attestations_traitees,
        })


logger = logging.getLogger(__name__)

@method_decorator([never_cache, ratelimit(key='user', rate='30/m', method='GET')], name='dispatch')
class DemanderInformationAPIView(APIView):
    """API pour envoyer une demande d'information complémentaire à un étudiant"""
    permission_classes = [IsAuthenticated]

    def post(self, request, demande_id):
        try:
            # Récupérer la demande
            demande = get_object_or_404(Demande, id=demande_id)
            
            # Récupérer le message de la demande
            message = request.data.get('message', '').strip()
            
            if not message:
                return Response({
                    'success': False,
                    'message': 'Le message est requis'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Vérifier que l'étudiant a un email
            if not demande.etudiant_email:
                return Response({
                    'success': False,
                    'message': 'Aucun email trouvé pour cet étudiant'
                }, status=status.HTTP_400_BAD_REQUEST)

            # ✅ ENVOI EMAIL SIMPLIFIÉ via le service
            email_envoye = self.envoyer_demande_information(demande, message)

            # Enregistrer l'action dans l'historique
            UserAction.objects.create(
                user=request.user,
                action=f"Demande d'information envoyée à {demande.etudiant_prenom} {demande.etudiant_nom}",
                performed_by=request.user
            )

            # 🔔 Notification push
            self.notifier_demande_information(demande, request.user, message)

            return Response({
                'success': True,
                'message': "Demande d'information envoyée avec succès",
                'email_envoye': email_envoye,
                'destinataire': {
                    'nom': f"{demande.etudiant_prenom} {demande.etudiant_nom}",
                    'email': demande.etudiant_email
                }
            }, status=status.HTTP_200_OK)

        except Demande.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Demande non trouvée'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Erreur lors de l'envoi de la demande d'information: {str(e)}")
            return Response({
                'success': False,
                'message': f"Erreur lors de l'envoi: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def envoyer_demande_information(self, demande, message):
        """
        ✅ MÉTHODE SIMPLIFIÉE : Envoie l'email via le service
        Tout le HTML a été déplacé dans le service email
        """
        try:
            # Construire l'URL de suivi
            base_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:8080')
            suivi_url = f"{base_url}/track-application"
            
            logger.info(f"📧 Envoi demande information à {demande.etudiant_email}")
            logger.info(f"🔗 URL de suivi: {suivi_url}")
            
            # ✅ UTILISER LE SERVICE EMAIL
            success = EmailSenderService.send_demande_information(
                demande=demande,
                message=message,
                suivi_url=suivi_url,
                async_send=True  # Envoi asynchrone
            )
            
            if success:
                logger.info(f"✅ Demande information envoyée à {demande.etudiant_email}")
            else:
                logger.warning(f"⚠️ Échec envoi demande information à {demande.etudiant_email}")
            
            return success
            
        except Exception as e:
            logger.error(f"❌ Erreur envoi demande information: {e}")
            return False

    def notifier_demande_information(self, demande, utilisateur, message):
        """Notifier la demande d'information via NotificationService"""
        try:
            NotificationService.notifier_demande_information(
                demande=demande,
                utilisateur=utilisateur,
                message=message
            )
            
            logger.info(f"✅ Notification push envoyée pour {utilisateur.email}")
            
        except Exception as e:
            logger.error(f"❌ Erreur notification demande information: {e}")

    
from services.rapport_service import RapportGeneratorService
from .models import PlanificationRapport
@method_decorator(ratelimit(key='user', rate='30/m', method='GET'), name='dispatch')
class ExportRapportAPIView(APIView):
    permission_classes = [IsAuthenticated, HasPermission(PERM_DATA_EXPORT)]

    RAPPORT_HANDLERS = {
        # Opérationnel
        'demandes_en_cours':        'generer_demandes_en_cours',
        'stages_actifs':            'generer_stages_actifs',
        'fins_imminentes':          'generer_fins_imminentes',
        'attestations_en_attente':  'generer_attestations_en_attente',
        # Performance
        'synthese_mensuelle':       'generer_synthese_mensuelle',
        'rapport_annuel':           'generer_rapport_annuel',
        'performance_traitement':   'generer_performance_traitement',
        # RH
        'repartition_directions':   'generer_repartition_directions',
        'profil_stagiaires':        'generer_profil_stagiaires',
        'etablissements_partenaires':'generer_etablissements_partenaires',
        'renouvellements':          'generer_renouvellements',
        # Conformité
        'audit_actions':            'generer_audit_actions',
        'demandes_archivees':       'generer_demandes_archivees',
    }

    def get(self, request):
        type_rapport = request.query_params.get('type')
        format_doc   = request.query_params.get('format_doc', 'pdf')  # pdf | excel

        if not type_rapport or type_rapport not in self.RAPPORT_HANDLERS:
            return Response({'error': 'Type de rapport invalide.'}, status=400)

        handler_name = self.RAPPORT_HANDLERS[type_rapport]
        handler = getattr(RapportGeneratorService, handler_name)

        try:
            params = dict(request.query_params)
            params.pop('type', None)
            params.pop('format_doc', None)
            # Aplatir les listes (QueryDict retourne des listes)
            params = {k: v[0] if isinstance(v, list) else v for k, v in params.items()}

            if format_doc == 'excel':
                buffer, filename = handler(params, format='excel')
                content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                ext = 'xlsx'
            else:
                buffer, filename = handler(params, format='pdf')
                content_type = 'application/pdf'
                ext = 'pdf'

            response = HttpResponse(buffer.getvalue(), content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}.{ext}"'
            return response

        except Exception as e:
            logger.error(f"Erreur génération rapport {type_rapport}: {e}", exc_info=True)
            return Response({'error': str(e)}, status=500)
        
